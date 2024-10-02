from openpyxl import load_workbook
import os

# Carregar a planilha modelo e as planilhas de origem
print("Carregando planilhas...")
modelo_wb = load_workbook(r'C:\\Users\\s-Lucas.SAraujo\\Downloads\\modelo.xlsx')
workbook1 = load_workbook(r'C:\\Users\\s-Lucas.SAraujo\\Downloads\\exportacaojeceaba.xlsx')
workbook2 = load_workbook(r'C:\\Users\\s-Lucas.SAraujo\\Downloads\\exportacaobarreiro.xlsx')

# Selecionar as abas ativas (pode alterar se for uma aba específica)
modelo_sheet = modelo_wb.active
sheet1 = workbook1.active
sheet2 = workbook2.active

# Mapeamento das colunas do modelo para as colunas das planilhas de origem
column_mapping = {
    'Data': ['Data'],
    'Identificador': ['Identificador'],
    'Classe': ['Classe'],
    'Turno': ['Turno'],
    'Hora inicial': ['Hora inicial'],
    'Hora final': ['Hora final'],
    'Tempo de intervalo': ['Tempo de intervalo'],
    'Área': ['Área']
}

# Função para encontrar o índice da coluna pelo cabeçalho
def get_column_index(sheet, header_name):
    for col in sheet.iter_cols(1, sheet.max_column):
        for cell in col:
            if cell.value == header_name:
                return cell.column
    return None

# Função para copiar dados de acordo com o mapeamento
def copy_data(source_sheet, target_sheet, mapping):
    print(f"Copiando dados da planilha {source_sheet.title}...")

    # Obter índices das colunas de destino
    target_indices = {target_column: get_column_index(target_sheet, target_column) for target_column in mapping}

    # Obter índices das colunas de origem
    source_indices = {}
    for target_column, source_columns in mapping.items():
        for source_column_name in source_columns:
            source_index = get_column_index(source_sheet, source_column_name)
            if source_index:
                source_indices[source_column_name] = source_index
                break

    # Armazenar dados temporariamente
    data_to_copy = []
    for row in source_sheet.iter_rows(min_row=3, max_row=source_sheet.max_row):  # Pular as duas primeiras linhas
        row_data = {}
        for target_column, source_columns in mapping.items():
            for source_column_name in source_columns:
                if source_column_name in source_indices:
                    row_data[target_indices[target_column]] = row[source_indices[source_column_name] - 1].value
                    break
        data_to_copy.append(row_data)

    # Escrever dados na planilha de destino
    for row_data in data_to_copy:
        target_row = target_sheet.max_row + 1
        for col_index, value in row_data.items():
            target_sheet.cell(row=target_row, column=col_index).value = value

# Copiar dados da primeira planilha para a planilha modelo
copy_data(sheet1, modelo_sheet, column_mapping)

# Função para alterar dados em massa
def alterar_dados_em_massa(sheet, coluna, valor_antigo, valor_novo):
    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row):  # Pular as duas primeiras linhas
        for cell in row:
            if cell.column_letter == coluna and cell.value == valor_antigo:
                cell.value = valor_novo

# Exemplo: alterar todos os valores "Antigo" na coluna 'B' para "Novo"
alterar_dados_em_massa(sheet1, 'B', 'Antigo', 'Novo')

# Copiar dados da segunda planilha para a planilha modelo
copy_data(sheet2, modelo_sheet, column_mapping)

# Salvar a planilha combinada
print("Salvando a planilha combinada...")
modelo_wb.save(r'C:\\Users\\s-Lucas.SAraujo\\Downloads\\planilha_combinada.xlsx')

# Abrir o arquivo automaticamente
print("Abrindo a planilha combinada...")
os.startfile(r'C:\\Users\\s-Lucas.SAraujo\\Downloads\\planilha_combinada.xlsx')
print("Processo concluído!")
