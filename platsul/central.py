import PySimpleGUI as sg
from openpyxl import load_workbook

class TelaEscoamento:

    def __init__(self):

        local = ["OCTG - PÁTIOCRL","OCTG - PT1ZL45","OCTG - PT1TTM","OCTG - PT1EMAG","OCTG - TLHCR","OCTG - NOV","VTI - QUADRATURA","VTI - FORJA","VTI - PÓRTICO","VTI - TT2","RK - RT3","RK - RT3","RK - RT4LINHA","RK - RT4LINHA","RK - RT5LINHA","RK - RT4GALPÃO","RK - RT4GALPÃO 2","RK - RT5GALPÃO 2","RK - RL1","RK - RL2","LA - RT4","LA - RT4","LA - RT5","LA - RT6","LA - RT6","LA - RT7","LA - TLHPA","LA - TLHPA","LA - RL6","LA - RL7","LA - RL8","RK - ST1","RK - ST1","RK - ST2","RK - ST3"]
        
        destino = ["ACIARIA SUCATA", "BALANÇA ACIARIA", "BALANÇA DO CARVÃO", "BIP", "BREDERO", "BREDERO PATIO-1", "BREDERO PATIO-2", "CANTEIRO TRADIMAQ 1", "CANTEIRO TRADIMAQ 2", "CENTRAL", "CL1", "CL2", "CL3", "CL4", "CL4 TENDA", "COLAPSO", "CRA", "CRP", "CVF/Q", "DPA", "DL1", "DL2", "DL3", "EQUIPE DESCARGA", "ESCORIAL", "FORJA MANUTENÇÃO", "FORJA MANUTENÇÃO GALPÃO", "FORJA MATÉRIA PRIMA", "FORJA MATÉRIA PRIMA GALPÃO", "GALPÃOTLP", "GALPÃOTT1", "LL1PC", "LL2PC", "LOCAL 34", "LOCAL 35", "LOCAL40", "LUVAS", "OFCENTRAL", "PÁTIO BAIXO FORNO", "PÁTIO CALÇADO", "PÁTIO CARVÃO", "PÁTIO DE ASFALTO", "PATIOCRL", "PATIOCRP", "PATIOFORJA", "PATIOLL2PC", "PATIORL4", "PATIOST1", "PATIOST1 TENDA", "PATIOTLHCR", "PLATAFORMA NORTE", "PLATSUL", "PLATSUL VTI", "PORTICO TRIAGEM", "PORTICOCROMO", "PS CONTINUO", "PT1CASING", "PT1EMAG", "PT1LUVAS", "PT1TTM", "PT1ZL45", "RL1", "RL2", "RL4", "RL6", "RL7", "RL8", "RL8_VD", "RODOMATICO", "RT1PC", "RT2PC", "RT3PC", "RT4GALPÃO", "RT4GALPÃO2", "RT4LINHA", "RT4PA", "RT5GALPÃO", "RT5GALPÃO2", "RT5LINHA", "RT5PA", "RT6PA", "RT7PA", "SERRASVAGNER-PC", "SINTERIZAÇÃO", "SL2FRENTE", "SL2MEIO", "ST1", "ST2", "ST3", "TLB", "TLD", "TLE", "TLF", "TLGCV", "TLGPA", "TLHCR", "TLHPA", "TT2", "VTS", "X2"]

        sg.theme('Reddit')

        Origem = [
            [sg.Text("Origem",size=(15,0)), sg.Text("T1"), sg.Text("T2"), sg.Text("T3"), sg.Text("Destino", size=(28,0)), sg.Text("Cliente", size=(9,0)), sg.Text("Descrição do Material"), sg.Text("Volume", size=(10,0)), sg.Text("Ordem de venda/item"), sg.Text("Prazo de Expedição"), sg.Text("Tempo Médio Carregamento"), sg.Text("Qtd. Carretas/Hora")],
            [sg.Text(local[0],size=(15,0)), sg.Input(size=(2,0), key='0'), sg.Input(size=(2,0), key='35'), sg.Input(size=(2,0), key='70'), sg.Combo(destino,size=(30,1), key='105'), sg.In(size=(10,0), key='140'), sg.In(size=(19,0), key='175'), sg.In(size=(11,0), key='210'), sg.In(size=(17,0), key='245'), sg.In(size=(17,0), key='280'), sg.In(size=(24,0), key='315'), sg.In(size=(15,0), key='350')],
            [sg.Text(local[1],size=(15,0)), sg.Input(size=(2,0), key='1'), sg.Input(size=(2,0), key='36'), sg.Input(size=(2,0), key='71'), sg.Combo(destino,size=(30,1), key='106'), sg.In(size=(10,0), key='141'), sg.In(size=(19,0), key='176'), sg.In(size=(11,0), key='211'), sg.In(size=(17,0), key='246'), sg.In(size=(17,0), key='281'), sg.In(size=(24,0), key='316'), sg.In(size=(15,0), key='351')],
            [sg.Text(local[2],size=(15,0)), sg.Input(size=(2,0), key='2'), sg.Input(size=(2,0), key='37'), sg.Input(size=(2,0), key='72'), sg.Combo(destino,size=(30,1), key='107'), sg.In(size=(10,0), key='142'), sg.In(size=(19,0), key='177'), sg.In(size=(11,0), key='212'), sg.In(size=(17,0), key='247'), sg.In(size=(17,0), key='282'), sg.In(size=(24,0), key='317'), sg.In(size=(15,0), key='352')],
            [sg.Text(local[3],size=(15,0)), sg.Input(size=(2,0), key='3'), sg.Input(size=(2,0), key='38'), sg.Input(size=(2,0), key='73'), sg.Combo(destino,size=(30,1), key='108'), sg.In(size=(10,0), key='143'), sg.In(size=(19,0), key='178'), sg.In(size=(11,0), key='213'), sg.In(size=(17,0), key='248'), sg.In(size=(17,0), key='283'), sg.In(size=(24,0), key='318'), sg.In(size=(15,0), key='353')],
            [sg.Text(local[4],size=(15,0)), sg.Input(size=(2,0), key='4'), sg.Input(size=(2,0), key='39'), sg.Input(size=(2,0), key='74'), sg.Combo(destino,size=(30,1), key='109'), sg.In(size=(10,0), key='144'), sg.In(size=(19,0), key='179'), sg.In(size=(11,0), key='214'), sg.In(size=(17,0), key='249'), sg.In(size=(17,0), key='284'), sg.In(size=(24,0), key='319'), sg.In(size=(15,0), key='354')],
            [sg.Text(local[5],size=(15,0)), sg.Input(size=(2,0), key='5'), sg.Input(size=(2,0), key='40'), sg.Input(size=(2,0), key='75'), sg.Combo(destino,size=(30,1), key='110'), sg.In(size=(10,0), key='145'), sg.In(size=(19,0), key='180'), sg.In(size=(11,0), key='215'), sg.In(size=(17,0), key='250'), sg.In(size=(17,0), key='285'), sg.In(size=(24,0), key='320'), sg.In(size=(15,0), key='355')],
            [sg.Text(local[6],size=(15,0)), sg.Input(size=(2,0), key='6'), sg.Input(size=(2,0), key='41'), sg.Input(size=(2,0), key='76'), sg.Combo(destino,size=(30,1), key='111'), sg.In(size=(10,0), key='146'), sg.In(size=(19,0), key='181'), sg.In(size=(11,0), key='216'), sg.In(size=(17,0), key='251'), sg.In(size=(17,0), key='286'), sg.In(size=(24,0), key='321'), sg.In(size=(15,0), key='356')],
            [sg.Text(local[7],size=(15,0)), sg.Input(size=(2,0), key='7'), sg.Input(size=(2,0), key='42'), sg.Input(size=(2,0), key='77'), sg.Combo(destino,size=(30,1), key='112'), sg.In(size=(10,0), key='147'), sg.In(size=(19,0), key='182'), sg.In(size=(11,0), key='217'), sg.In(size=(17,0), key='252'), sg.In(size=(17,0), key='287'), sg.In(size=(24,0), key='322'), sg.In(size=(15,0), key='357')],
            [sg.Text(local[8],size=(15,0)), sg.Input(size=(2,0), key='8'), sg.Input(size=(2,0), key='43'), sg.Input(size=(2,0), key='78'), sg.Combo(destino,size=(30,1), key='113'), sg.In(size=(10,0), key='148'), sg.In(size=(19,0), key='183'), sg.In(size=(11,0), key='218'), sg.In(size=(17,0), key='253'), sg.In(size=(17,0), key='288'), sg.In(size=(24,0), key='323'), sg.In(size=(15,0), key='358')],
            [sg.Text(local[9],size=(15,0)), sg.Input(size=(2,0), key='9'), sg.Input(size=(2,0), key='44'), sg.Input(size=(2,0), key='79'), sg.Combo(destino,size=(30,1), key='114'), sg.In(size=(10,0), key='149'), sg.In(size=(19,0), key='184'), sg.In(size=(11,0), key='219'), sg.In(size=(17,0), key='254'), sg.In(size=(17,0), key='289'), sg.In(size=(24,0), key='324'), sg.In(size=(15,0), key='359')],
            [sg.Text(local[10],size=(15,0)), sg.Input(size=(2,0), key='10'), sg.Input(size=(2,0), key='45'), sg.Input(size=(2,0), key='80'), sg.Combo(destino,size=(30,1), key='115'), sg.In(size=(10,0), key='150'), sg.In(size=(19,0), key='185'), sg.In(size=(11,0), key='220'), sg.In(size=(17,0), key='255'), sg.In(size=(17,0), key='290'), sg.In(size=(24,0), key='325'), sg.In(size=(15,0), key='360')],
            [sg.Text(local[11],size=(15,0)), sg.Input(size=(2,0), key='11'), sg.Input(size=(2,0), key='46'), sg.Input(size=(2,0), key='81'), sg.Combo(destino,size=(30,1), key='116'), sg.In(size=(10,0), key='151'), sg.In(size=(19,0), key='186'), sg.In(size=(11,0), key='221'), sg.In(size=(17,0), key='256'), sg.In(size=(17,0), key='291'), sg.In(size=(24,0), key='326'), sg.In(size=(15,0), key='361')],
            [sg.Text(local[12],size=(15,0)), sg.Input(size=(2,0), key='12'), sg.Input(size=(2,0), key='47'), sg.Input(size=(2,0), key='82'), sg.Combo(destino,size=(30,1), key='117'), sg.In(size=(10,0), key='152'), sg.In(size=(19,0), key='187'), sg.In(size=(11,0), key='222'), sg.In(size=(17,0), key='257'), sg.In(size=(17,0), key='292'), sg.In(size=(24,0), key='327'), sg.In(size=(15,0), key='362')],
            [sg.Text(local[13],size=(15,0)), sg.Input(size=(2,0), key='13'), sg.Input(size=(2,0), key='48'), sg.Input(size=(2,0), key='83'), sg.Combo(destino,size=(30,1), key='118'), sg.In(size=(10,0), key='153'), sg.In(size=(19,0), key='188'), sg.In(size=(11,0), key='223'), sg.In(size=(17,0), key='258'), sg.In(size=(17,0), key='293'), sg.In(size=(24,0), key='328'), sg.In(size=(15,0), key='363')],
            [sg.Text(local[14],size=(15,0)), sg.Input(size=(2,0), key='14'), sg.Input(size=(2,0), key='49'), sg.Input(size=(2,0), key='84'), sg.Combo(destino,size=(30,1), key='119'), sg.In(size=(10,0), key='154'), sg.In(size=(19,0), key='189'), sg.In(size=(11,0), key='224'), sg.In(size=(17,0), key='259'), sg.In(size=(17,0), key='294'), sg.In(size=(24,0), key='329'), sg.In(size=(15,0), key='364')],
            [sg.Text(local[15],size=(15,0)), sg.Input(size=(2,0), key='15'), sg.Input(size=(2,0), key='50'), sg.Input(size=(2,0), key='85'), sg.Combo(destino,size=(30,1), key='120'), sg.In(size=(10,0), key='155'), sg.In(size=(19,0), key='190'), sg.In(size=(11,0), key='225'), sg.In(size=(17,0), key='260'), sg.In(size=(17,0), key='295'), sg.In(size=(24,0), key='330'), sg.In(size=(15,0), key='365')],
            [sg.Text(local[16],size=(15,0)), sg.Input(size=(2,0), key='16'), sg.Input(size=(2,0), key='51'), sg.Input(size=(2,0), key='86'), sg.Combo(destino,size=(30,1), key='121'), sg.In(size=(10,0), key='156'), sg.In(size=(19,0), key='191'), sg.In(size=(11,0), key='226'), sg.In(size=(17,0), key='261'), sg.In(size=(17,0), key='296'), sg.In(size=(24,0), key='331'), sg.In(size=(15,0), key='366')],
            [sg.Text(local[17],size=(15,0)), sg.Input(size=(2,0), key='17'), sg.Input(size=(2,0), key='52'), sg.Input(size=(2,0), key='87'), sg.Combo(destino,size=(30,1), key='122'), sg.In(size=(10,0), key='157'), sg.In(size=(19,0), key='192'), sg.In(size=(11,0), key='227'), sg.In(size=(17,0), key='262'), sg.In(size=(17,0), key='297'), sg.In(size=(24,0), key='332'), sg.In(size=(15,0), key='367')],
            [sg.Text(local[18],size=(15,0)), sg.Input(size=(2,0), key='18'), sg.Input(size=(2,0), key='53'), sg.Input(size=(2,0), key='88'), sg.Combo(destino,size=(30,1), key='123'), sg.In(size=(10,0), key='158'), sg.In(size=(19,0), key='193'), sg.In(size=(11,0), key='228'), sg.In(size=(17,0), key='263'), sg.In(size=(17,0), key='298'), sg.In(size=(24,0), key='333'), sg.In(size=(15,0), key='368')],
            [sg.Text(local[19],size=(15,0)), sg.Input(size=(2,0), key='19'), sg.Input(size=(2,0), key='54'), sg.Input(size=(2,0), key='89'), sg.Combo(destino,size=(30,1), key='124'), sg.In(size=(10,0), key='159'), sg.In(size=(19,0), key='194'), sg.In(size=(11,0), key='229'), sg.In(size=(17,0), key='264'), sg.In(size=(17,0), key='299'), sg.In(size=(24,0), key='334'), sg.In(size=(15,0), key='369')],
            [sg.Text(local[20],size=(15,0)), sg.Input(size=(2,0), key='20'), sg.Input(size=(2,0), key='55'), sg.Input(size=(2,0), key='90'), sg.Combo(destino,size=(30,1), key='125'), sg.In(size=(10,0), key='160'), sg.In(size=(19,0), key='195'), sg.In(size=(11,0), key='230'), sg.In(size=(17,0), key='265'), sg.In(size=(17,0), key='300'), sg.In(size=(24,0), key='335'), sg.In(size=(15,0), key='370')],
            [sg.Text(local[21],size=(15,0)), sg.Input(size=(2,0), key='21'), sg.Input(size=(2,0), key='56'), sg.Input(size=(2,0), key='91'), sg.Combo(destino,size=(30,1), key='126'), sg.In(size=(10,0), key='161'), sg.In(size=(19,0), key='196'), sg.In(size=(11,0), key='231'), sg.In(size=(17,0), key='266'), sg.In(size=(17,0), key='301'), sg.In(size=(24,0), key='336'), sg.In(size=(15,0), key='371')],
            [sg.Text(local[22],size=(15,0)), sg.Input(size=(2,0), key='22'), sg.Input(size=(2,0), key='57'), sg.Input(size=(2,0), key='92'), sg.Combo(destino,size=(30,1), key='127'), sg.In(size=(10,0), key='162'), sg.In(size=(19,0), key='197'), sg.In(size=(11,0), key='232'), sg.In(size=(17,0), key='267'), sg.In(size=(17,0), key='302'), sg.In(size=(24,0), key='337'), sg.In(size=(15,0), key='372')],
            [sg.Text(local[23],size=(15,0)), sg.Input(size=(2,0), key='23'), sg.Input(size=(2,0), key='58'), sg.Input(size=(2,0), key='93'), sg.Combo(destino,size=(30,1), key='128'), sg.In(size=(10,0), key='163'), sg.In(size=(19,0), key='198'), sg.In(size=(11,0), key='233'), sg.In(size=(17,0), key='268'), sg.In(size=(17,0), key='303'), sg.In(size=(24,0), key='338'), sg.In(size=(15,0), key='373')],
            [sg.Text(local[24],size=(15,0)), sg.Input(size=(2,0), key='24'), sg.Input(size=(2,0), key='59'), sg.Input(size=(2,0), key='94'), sg.Combo(destino,size=(30,1), key='129'), sg.In(size=(10,0), key='164'), sg.In(size=(19,0), key='199'), sg.In(size=(11,0), key='234'), sg.In(size=(17,0), key='269'), sg.In(size=(17,0), key='304'), sg.In(size=(24,0), key='339'), sg.In(size=(15,0), key='374')],
            [sg.Text(local[25],size=(15,0)), sg.Input(size=(2,0), key='25'), sg.Input(size=(2,0), key='60'), sg.Input(size=(2,0), key='95'), sg.Combo(destino,size=(30,1), key='130'), sg.In(size=(10,0), key='165'), sg.In(size=(19,0), key='200'), sg.In(size=(11,0), key='235'), sg.In(size=(17,0), key='270'), sg.In(size=(17,0), key='305'), sg.In(size=(24,0), key='340'), sg.In(size=(15,0), key='375')],
            [sg.Text(local[26],size=(15,0)), sg.Input(size=(2,0), key='26'), sg.Input(size=(2,0), key='61'), sg.Input(size=(2,0), key='96'), sg.Combo(destino,size=(30,1), key='131'), sg.In(size=(10,0), key='166'), sg.In(size=(19,0), key='201'), sg.In(size=(11,0), key='236'), sg.In(size=(17,0), key='271'), sg.In(size=(17,0), key='306'), sg.In(size=(24,0), key='341'), sg.In(size=(15,0), key='376')],
            [sg.Text(local[27],size=(15,0)), sg.Input(size=(2,0), key='27'), sg.Input(size=(2,0), key='62'), sg.Input(size=(2,0), key='97'), sg.Combo(destino,size=(30,1), key='132'), sg.In(size=(10,0), key='167'), sg.In(size=(19,0), key='202'), sg.In(size=(11,0), key='237'), sg.In(size=(17,0), key='272'), sg.In(size=(17,0), key='307'), sg.In(size=(24,0), key='342'), sg.In(size=(15,0), key='377')],
            [sg.Text(local[28],size=(15,0)), sg.Input(size=(2,0), key='28'), sg.Input(size=(2,0), key='63'), sg.Input(size=(2,0), key='98'), sg.Combo(destino,size=(30,1), key='133'), sg.In(size=(10,0), key='168'), sg.In(size=(19,0), key='203'), sg.In(size=(11,0), key='238'), sg.In(size=(17,0), key='273'), sg.In(size=(17,0), key='308'), sg.In(size=(24,0), key='343'), sg.In(size=(15,0), key='378')],
            [sg.Text(local[29],size=(15,0)), sg.Input(size=(2,0), key='29'), sg.Input(size=(2,0), key='64'), sg.Input(size=(2,0), key='99'), sg.Combo(destino,size=(30,1), key='134'), sg.In(size=(10,0), key='169'), sg.In(size=(19,0), key='204'), sg.In(size=(11,0), key='239'), sg.In(size=(17,0), key='274'), sg.In(size=(17,0), key='309'), sg.In(size=(24,0), key='344'), sg.In(size=(15,0), key='379')],
            [sg.Text(local[30],size=(15,0)), sg.Input(size=(2,0), key='30'), sg.Input(size=(2,0), key='65'), sg.Input(size=(2,0), key='100'), sg.Combo(destino,size=(30,1), key='135'), sg.In(size=(10,0), key='170'), sg.In(size=(19,0), key='205'), sg.In(size=(11,0), key='240'), sg.In(size=(17,0), key='275'), sg.In(size=(17,0), key='310'), sg.In(size=(24,0), key='345'), sg.In(size=(15,0), key='380')],
            [sg.Text(local[31],size=(15,0)), sg.Input(size=(2,0), key='31'), sg.Input(size=(2,0), key='66'), sg.Input(size=(2,0), key='101'), sg.Combo(destino,size=(30,1), key='136'), sg.In(size=(10,0), key='171'), sg.In(size=(19,0), key='206'), sg.In(size=(11,0), key='241'), sg.In(size=(17,0), key='276'), sg.In(size=(17,0), key='311'), sg.In(size=(24,0), key='346'), sg.In(size=(15,0), key='381')],
            [sg.Text(local[32],size=(15,0)), sg.Input(size=(2,0), key='32'), sg.Input(size=(2,0), key='67'), sg.Input(size=(2,0), key='102'), sg.Combo(destino,size=(30,1), key='137'), sg.In(size=(10,0), key='172'), sg.In(size=(19,0), key='207'), sg.In(size=(11,0), key='242'), sg.In(size=(17,0), key='277'), sg.In(size=(17,0), key='312'), sg.In(size=(24,0), key='347'), sg.In(size=(15,0), key='382')],
            [sg.Text(local[33],size=(15,0)), sg.Input(size=(2,0), key='33'), sg.Input(size=(2,0), key='68'), sg.Input(size=(2,0), key='103'), sg.Combo(destino,size=(30,1), key='138'), sg.In(size=(10,0), key='173'), sg.In(size=(19,0), key='208'), sg.In(size=(11,0), key='243'), sg.In(size=(17,0), key='278'), sg.In(size=(17,0), key='313'), sg.In(size=(24,0), key='348'), sg.In(size=(15,0), key='383')],
            [sg.Text(local[34],size=(15,0)), sg.Input(size=(2,0), key='34'), sg.Input(size=(2,0), key='69'), sg.Input(size=(2,0), key='104'), sg.Combo(destino,size=(30,1), key='139'), sg.In(size=(10,0), key='174'), sg.In(size=(19,0), key='209'), sg.In(size=(11,0), key='244'), sg.In(size=(17,0), key='279'), sg.In(size=(17,0), key='314'), sg.In(size=(24,0), key='349'), sg.In(size=(15,0), key='384')],
            [sg.Combo(local,size=(15,0), key='op4'), sg.Input(size=(2,0), key='op0'), sg.Input(size=(2,0), key='op1'), sg.Input(size=(2,0), key='op2'), sg.Combo(destino,size=(30,1), key='op3'), sg.In(size=(10,0), key='op5'), sg.In(size=(19,0), key='op6'), sg.In(size=(11,0), key='op7'), sg.In(size=(17,0), key='op8'), sg.In(size=(17,0), key='op9'), sg.In(size=(24,0), key='op10'), sg.In(size=(15,0), key='op11')],
            [sg.Text("Data (dd/mm/yyyy):",size=(12,0)), sg.In(key="-CAL-", enable_events=True, font=('Arial', 11), size=(17,0)), sg.CalendarButton('Data', key='calendario', format=('%d/%m/%Y'), size=(6,0), font=('Arial', 10))],
            [sg.Button("Enviar dados" ,font=('Arial', 12)), sg.Quit(font=('Arial', 12)), sg.Button("Limpar" ,font=('Arial', 12)), sg.Button("Enviar dado Customizado" ,font=('Arial', 12))]
        ]

        layout = [
            [
            sg.Column(Origem, scrollable=True, vertical_scroll_only = True, expand_x=True, expand_y=True)
            ]
        ]

        self.janela = sg.Window('Plano de Escoamento', layout, resizable = True).Finalize()

        while True:

            self.button, self.values = self.janela.read()

            if self.button == "Limpar":
                for i in range(0, 385):
                    self.janela[str(i)]('')

            if self.button == "Enviar dado Customizado":
                aux = 0
                for i in range(0, 12):
                    text = self.values["op"+str(i)]
                    if text == '':
                        aux += 1
                    else:
                        pass
                
                text2 = self.values["-CAL-"]
                if aux != 10 and text2 != '':
                    wb = load_workbook(r'C:\Users\felipe.malves\Vallourec\Global Data Factory - 01. NLI001 -  Plano D+7 Plataforma Sul\Plano Semanal Escoamento.xlsx')
                    ws = wb['PlanoEscoamento']
                    max_row = ws.max_row
                    for i in range(0, 1):
                        ws.cell(max_row+1, 1).value = self.values['-CAL-']
                        ws.cell(max_row+1, 4).value = self.values["op"+str(i)]
                        ws.cell(max_row+1, 5).value = self.values["op"+str(1+i)]
                        ws.cell(max_row+1, 6).value = self.values["op"+str(2+i)]
                        ws.cell(max_row+1, 3).value = self.values["op"+str(3+i)]
                        ws.cell(max_row+1, 2).value = self.values["op"+str(4+i)]
                        ws.cell(max_row+1, 7).value = self.values["op"+str(5+i)]
                        ws.cell(max_row+1, 8).value = self.values["op"+str(6+i)]
                        ws.cell(max_row+1, 9).value = self.values["op"+str(7+i)]
                        ws.cell(max_row+1, 10).value = self.values["op"+str(8+i)]
                        ws.cell(max_row+1, 11).value = self.values["op"+str(9+i)]
                        ws.cell(max_row+1, 11).value = self.values["op"+str(10+i)]
                        ws.cell(max_row+1, 11).value = self.values["op"+str(11+i)]
                        max_row = ws.max_row
                    table = ws.tables["TabelaEscoamento"]
                    table.ref = "A1:M"+str(max_row)
                    wb.save(r'C:\Users\felipe.malves\Vallourec\Global Data Factory - 01. NLI001 -  Plano D+7 Plataforma Sul\Plano Semanal Escoamento.xlsx')
                    sg.popup("Valores enviados!")
                    for i in range(0, 12):
                        self.janela["op"+str(i)]('')
                else:
                    sg.popup_error("Preencha algum campo!")

            if self.button in (sg.WINDOW_CLOSED, "Quit"):
                break
            elif self.button == "Enviar dados":
                text = self.values["-CAL-"]
                if text != '':
                    wb = load_workbook(r'C:\Users\felipe.malves\Vallourec\Global Data Factory - 01. NLI001 -  Plano D+7 Plataforma Sul\Plano Semanal Escoamento.xlsx')
                    ws = wb['PlanoEscoamento']
                    max_row = ws.max_row
                    for i in range(0, 35):
                        ws.cell(max_row+1, 1).value = self.values['-CAL-']
                        ws.cell(max_row+1, 4).value = self.values[str(i)]
                        ws.cell(max_row+1, 5).value = self.values[str(35+i)]
                        ws.cell(max_row+1, 6).value = self.values[str(70+i)]
                        ws.cell(max_row+1, 3).value = self.values[str(105+i)]
                        ws.cell(max_row+1, 7).value = self.values[str(140+i)]
                        ws.cell(max_row+1, 8).value = self.values[str(175+i)]
                        ws.cell(max_row+1, 9).value = self.values[str(210+i)]
                        ws.cell(max_row+1, 10).value = self.values[str(245+i)]
                        ws.cell(max_row+1, 11).value = self.values[str(280+i)]
                        ws.cell(max_row+1, 12).value = self.values[str(315+i)]
                        ws.cell(max_row+1, 13).value = self.values[str(350+i)]
                        ws.cell(max_row+1, 2).value = local[i]
                        max_row = ws.max_row
                    i = 1
                    while i < max_row + 1:
                        aux=0
                        for j in range(1, 14):
                            if ws.cell(i, j).value == '':
                                aux += 1
                        if aux == 11:
                            ws.delete_rows(i, 1)
                            max_row = ws.max_row
                        else:
                            i += 1
                    table = ws.tables["TabelaEscoamento"]
                    table.ref = "A1:M"+str(max_row)
                    wb.save(r'C:\Users\felipe.malves\Vallourec\Global Data Factory - 01. NLI001 -  Plano D+7 Plataforma Sul\Plano Semanal Escoamento.xlsx')
                    sg.popup("Valores enviados!")
                else:
                    sg.popup_error('Escolha uma data!')

TelaEscoamento()