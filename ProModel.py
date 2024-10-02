from openpyxl import load_workbook
from suporte.Recursos import Loop
import os
import numpy as np
import pandas as pd

#PREPARANDO RECURSOS
user = os.path.expanduser('~')

BaseMath = np.array([["BALANÇA ACIARIA","BALANCA_ACIARIA","ACIARIA"],
["ACIARIA SUCATA","ACIARIA_SUCATA","ACIARIA"],
["AC04","AC04","ACIARIA"],
["RT5PA","RT5PA","AJUSTAGEM LA"],
["RT4PA","RT4PA","AJUSTAGEM LA"],
["RT6PA","RT6PA","AJUSTAGEM LA"],
["RODOMATICO","RODOMATICO","AJUSTAGEM LA"],
["TLHPA","TLHPA","AJUSTAGEM LA"],
["TLGPA","TLGPA","AJUSTAGEM LA"],
["PATIOTLGPA","TLGPA","AJUSTAGEM LA"],
["RT7PA","RT7PA","AJUSTAGEM LA"],
["PATIO_RT5PA","PATIO_RT5PA","AJUSTAGEM LA"],
["COLAPSO","COLAPSO","AJUSTAGEM LA"],
["RL2","RL2","AJUSTAGEM RK"],
["RT5LINHA","RT5LINHA","AJUSTAGEM RK"],
["RT4LINHA","RT4LINHA","AJUSTAGEM RK"],
["RL1","RL1","AJUSTAGEM RK"],
["RT3PC","RT3PC","AJUSTAGEM RK"],
["RT4GALPÃO2","RT4GALPAO2","AJUSTAGEM RK"],
["RT4GALPÃO","RT4GALPAO","AJUSTAGEM RK"],
["RT5GALPÃO","RT5GALPAO","AJUSTAGEM RK"],
["BREDERO PATIO-1","BREDERO_PATIO_1","BREDERO"],
["VTS","VTS","BREDERO"],
["BREDERO PATIO-2","BREDERO_PATIO_2","BREDERO"],
["PÁTIO CARVÃO","PATIO_CARVAO","CARVÃO"],
["BALANÇA DO CARVÃO","BALANCA_DO_CARVAO","CARVÃO"],
["CL2","CL2","DPA"],
["CL1","CL1","DPA"],
["CL3","CL3","DPA"],
["CL4","CL4","DPA"],
["DL2-BOX","DL2_BOX","DPA"],
["DL3-BEGE","DL3_BEGE","DPA"],
["DL3-LARANJA","DL3_LARANJA","DPA"],
["DL1RUA3","DL1RUA3","DPA"],
["DL1RUA7","DL1RUA7","DPA"],
["DL3-BRANCO","DL3_BRANCO","DPA"],
["DL1RUA9","DL1RUA9","DPA"],
["DL2-AMARELO","DL2_AMARELO","DPA"],
["DL1RUA6","DL1RUA6","DPA"],
["DL3-MARROM","DL3_MARROM","DPA"],
["DL3-VERDE","DL3_VERDE","DPA"],
["DL1RUA1","DL1RUA1","DPA"],
["DL2-AZUL","DL2_AZUL","DPA"],
["DL2-PRETO","DL2_PRETO","DPA"],
["DL1RUA4","DL1RUA4","DPA"],
["DL2-VERMELHO","DL2_VERMELHO","DPA"],
["DL1RUA10","DL1RUA10","DPA"],
["DL1RUA5","DL1RUA5","DPA"],
["DL1RUA2","DL1RUA2","DPA"],
["DL1RUA8","DL1RUA8","DPA"],
["DL1RUA8/SOBRECANAL","DL1RUA8","DPA"],
["DL1RUA10/SOBRECANAL","DL1RUA8","DPA"],
["DL1RUA9/SOBRECANAL","DL1RUA8","DPA"],
["DL1RUA6/SOBRECANAL","DL1RUA6", "DPA"],
["DL1RUA7/SOBRECANAL","DL1RUA7", "DPA"],
["CL4","CL4","DPA"],
["FORJA MATÉRIA PRIMA","FORJA_MATERIA_PRIMA","FORJA"],
["FORJA MANUTENÇÃO","FORJA_MANUTENCAO","FORJA"],
["RL4","RL4","LAMINADOR LA"],
["PATIORL4","PATIORL4","LAMINADOR LA"],
["RT2PC","RT2PC","LAMINADOR RK"],
["LL1PC","LL1PC","LAMINADOR RK"],
["LL2PC","LL2PC","LL2"],
["LUVAS","LUVAS","LUVAS"],
["OFCENTRAL","OFCENTRAL","OFCENTRAL"],
["OFCHAPA","OFCHAPA","OFCENTRAL"],
["OFSERRA","OFSERRA","OFCENTRAL"],
["OFCHAPA","OFCHAPA","OFCENTRAL"],
["OFPORTAO","OFPORTAO","OFCENTRAL"],
["OFJATO","OFJATO","OFCENTRAL"],
["PLATAFORMA NORTE","PLATAFORMA_NORTE","PLATAFORMA NORTE"],
["BALANÇA PLATAFORMA NORTE","PLATAFORMA_NORTE","PLATAFORMA NORTE"],
["PLATSUL PORTÃO","PLATSUL_PORTAO","PLATSUL"],
["PLATSUL VTI","PLATSUL_PORTAO","PLATSUL"],
["PLATSUL TENDA","PLATSUL_PORTAO","PLATSUL"],
["PLATSUL TÉRREO ","PLATSUL_PORTAO","PLATSUL"],
["PLATSUL LATERAL PVPL","PLATSUL_LATERAL_PVPL","PLATSUL"],
["PLATSUL RUA 22","PLATSUL_RUA_22","PLATSUL"],
["PLATSUL RUA 10","PLATSUL_RUA_10","PLATSUL"],
["PLATSUL RUA 18","PLATSUL_RUA_18","PLATSUL"],
["PLATSUL RUA 23","PLATSUL_RUA_23","PLATSUL"],
["PLATSUL RUA 12","PLATSUL_RUA_12","PLATSUL"],
["PLATSUL RUA 15","PLATSUL_RUA_15","PLATSUL"],
["PLATSUL RUA 20","PLATSUL_RUA_20","PLATSUL"],
["PLATSUL RUA 6","PLATSUL_RUA_6","PLATSUL"],
["PLATSUL RUA 21","PLATSUL_RUA_21","PLATSUL"],
["PLATSUL RUA 13","PLATSUL_RUA_13","PLATSUL"],
["PLATSUL RUA 19","PLATSUL_RUA_19","PLATSUL"],
["PLATSUL RUA 17","PLATSUL_RUA_17","PLATSUL"],
["PLATSUL RUA 14","PLATSUL_RUA_14","PLATSUL"],
["PLATSUL RUA 7","PLATSUL_RUA_7","PLATSUL"],
["PLATSUL RUA 11","PLATSUL_RUA_11","PLATSUL"],
["PLATSUL RUA 16","PLATSUL_RUA_16","PLATSUL"],
["PLATSUL RUA 24","PLATSUL_RUA_24","PLATSUL"],
["PLATSUL RUA 9","PLATSUL_RUA_9","PLATSUL"],
["PLATSUL RUA 8","PLATSUL_RUA_8","PLATSUL"],
["CVF/Q","CVF_Q","QUADRATURA"],
["PT1EMAG","PT1EMAG","ROSQUEAMENTO "],
["PT1TTM","PT1TTM","ROSQUEAMENTO "],
["PATIOCRL","PATIOCRL","ROSQUEAMENTO "],
["CRA","CRA","ROSQUEAMENTO "],
["PT1ZL45","PT1ZL45","ROSQUEAMENTO "],
["X2","X2","ROSQUEAMENTO "],
["PT1LUVAS","PT1LUVAS","ROSQUEAMENTO "],
["TLHCR","TLHCR","ROSQUEAMENTO "],
["SL2MEIO","SL2MEIO","ROSQUEAMENTO "],
["SL2FRENTE","SL2FRENTE","ROSQUEAMENTO "],
["RL7","RL7","TÊMPERA LA"],
["PATIORL6","PATIORL6","TÊMPERA LA"],
["PATIORL6 - 1","PATIORL6","TÊMPERA LA"],
["PATIORL6 - 2","PATIORL6","TÊMPERA LA"],
["PATIORL6 - 3","PATIORL6","TÊMPERA LA"],
["RL6","RL6","TÊMPERA LA"],
["RL8_VD","RL8_VD","TÊMPERA LA"],
["PATIORL8","PATIORL8","TÊMPERA LA"],
["PATIORL8_VD","PATIORL8_VD","TÊMPERA LA"],
["PATIORL7","PATIORL7","TÊMPERA LA"],
["PATIOST1","PATIOST1","TÊMPERA RK"],
["ST1CABINE","ST1CABINE","TÊMPERA RK"],
["ST3","ST3","TÊMPERA RK"],
["ST2","ST2","TÊMPERA RK"],
["ST1 GALPÃO","ST1","TÊMPERA RK"],
["ST1","ST1","TÊMPERA RK"],
["PATIOST3","PATIOST3","TÊMPERA RK"],
["TT2","TT2","VTI"],
["TLE","TLE","VTI"],
["GALPÃOTT1 ","GALPAOTT1","VTI"],
["PATIOTLE","PATIOTLE","VTI"],
["PATIOTT2","PATIOTT2","VTI"],
["PORTICOCROMO ","PORTICOCROMO","VTI"],
["PORTICO TRIAGEM","PORTICO_TRIAGEM","VTI"],
["TLGCV","TLGCV","VTI"],
["TLF","TLF","VTI"],
["GALPÃOTLP","GALPAOTLP","VTI"],
["PATIOCVF/Q","PATIOCVF_Q","VTI"],
["PATIOTLF","PATIOTLF","VTI"],
["PATIOTLGCV","PATIOTLGCV","VTI"],
["PATIOFORJA","PATIOFORJA","VTI"],
["DL3","DL3_BRANCO","DPA"],
["DL2","DL2_AZUL","DPA"],
["DL1","DL1RUA6","DPA"],
["DPA","DL1RUA7","DPA"],
["BREDERO","BREDERO_PATIO_2","BREDERO"],
["PLATSUL","PLATSUL_RUA_16","PLATSUL"],
["SINTERIZAÇÃO","PLATSUL_PORTAO","PLATSUL"],
["LOCAL40","RL8_VD","TÊMPERA LA"],
["CL4 TENDA","CL4","DPA"],
["PATIO RT5PA","PATIO_RT5PA","AJUSTAGEM LA"],
["PATIOLL2PC","LL2PC","LL2"],
["PATIO RT7PA","RT7PA","AJUSTAGEM LA"],
["RT5GALPÃO2","RT5GALPAO","AJUSTAGEM RK"],
["PATIO MIRANTE","PATIO_CARVAO","CARVÃO"],
["PATIOTLHCR","PATIOTLHCR","ROSQUEAMENTO "],
["PATIOST1 TENDA","PATIOST1","TÊMPERA RK"],
["VTS PLATAFORMA NORTE","PLATAFORMA_NORTE","PLATAFORMA NORTE"],
["BARRAS PLATAFORMANORTE","PLATAFORMA_NORTE","PLATAFORMA NORTE"],
["VTS TRIAGEM BH","VTS","BREDERO"],
["RT1PC","RT2PC","LAMINADOR RK"],
["DL3 RUA 1","DL3_BRANCO","DPA"],
["DL3 RUA 10","DL3_BRANCO","DPA"],
["DL3 RUA 15","DL3_BRANCO","DPA"],
["DL3 RUA 18","DL3_BRANCO","DPA"],
["DL3 RUA 16","DL3_BRANCO","DPA"],
["DL3 RUA 7","DL3_BRANCO","DPA"],
["DL3 RUA 11","DL3_BRANCO","DPA"],
["DL3 RUA 6","DL3_BRANCO","DPA"],
["DL3 RUA 17","DL3_BRANCO","DPA"],
["DL3 RUA 14","DL3_BRANCO","DPA"],
["DL3 RUA 12","DL3_BRANCO","DPA"],
["DL3 RUA 13","DL3_BRANCO","DPA"],
["DL3 RUA 9","DL3_BRANCO","DPA"],
["DL3 RUA 4","DL3_BRANCO","DPA"],
["DL3 RUA 5","DL3_BRANCO","DPA"],
["DL3 RUA 3","DL3_BRANCO","DPA"],
["DL3 RUA 8","DL3_BRANCO","DPA"],
["DL3 RUA 2","DL3_BRANCO","DPA"],
["PÁTIO CALÇADO","PÁTIO CALÇADO","AJUSTAGEM RK"],
["RL8","RL8","TÊMPERA LA"]
])

wb = load_workbook(filename=(user + '\\Downloads\\Relatorio_viagens_carretinhas_ruckers.xlsx'))

dest_filename = user + '\\Downloads\\Teste.xlsx'

ws = wb['Sheet1']

max_row_ws = ws.max_row
max_column_ws = ws.max_column

#DELETANDO COISA DESNECESSÁRIA
ws.delete_cols(1, 26)
ws.delete_cols(2, 1)
ws.delete_cols(3, 15)
ws.delete_rows(1, 3)

Loop.Deletar(1, 1, max_row_ws+1, ws, "CANTEIRO TRADIMAQ 1")
Loop.Deletar(1, 1, max_row_ws+1, ws, "CANTEIRO TRADIMAQ 2")
Loop.Deletar(2, 1, max_row_ws+1, ws, "CANTEIRO TRADIMAQ 1")
Loop.Deletar(2, 1, max_row_ws+1, ws, "CANTEIRO TRADIMAQ 2")

for i in range(2, max_row_ws+1):
    if ws.cell(i, 1).value is None:
        break
    aux = False
    for item in BaseMath:
        if ws.cell(i, 1).value == item[0]:
            ws.cell(i, 1).value = item[1]
            aux = True
            break
    if aux == False:
        ws.cell(i, 1).value = "---"

for i in range(2, max_row_ws+1):
    if ws.cell(i, 2).value is None:
        break
    aux = False
    for item in BaseMath:
        if ws.cell(i, 2).value == item[0]:
            ws.cell(i, 2).value = item[1]
            aux = True
            break
    if aux == False:
            ws.cell(i, 2).value = "---"

Loop.Deletar(1, 1, max_row_ws+1, ws, "---")
Loop.Deletar(2, 1, max_row_ws+1, ws, "---")

#INSERINDO VALOR NÚMERICO PARA CONTABILIZAR VIAGENS REPETIDAS
ws.cell(1, 3).value = "Qtd. Viagens"
for i in range(2, max_row_ws+1):
    if ws.cell(i, 1).value is None:
        break
    ws.cell(i, 3).value = 1

#INSERINDO MACRO ÁREAS
ws.insert_cols(1,2)

ws.cell(1,1).value = "Macro Área Origem"
ws.cell(1,2).value = "Macro Área Destino"

for i in range(2, max_row_ws+1):
    if ws.cell(i, 3).value is None:
        break
    aux = False
    for item in BaseMath:
        if ws.cell(i, 3).value == item[1]:
            ws.cell(i, 1).value = item[2]
            aux = True
            break
    if aux == False:
            ws.cell(i, 1).value = "---"

for i in range(2, max_row_ws+1):
    if ws.cell(i, 4).value is None:
        break
    aux = False
    for item in BaseMath:
        if ws.cell(i, 4).value == item[1]:
            ws.cell(i, 2).value = item[2]
            aux = True
            break
    if aux == False:
            ws.cell(i, 2).value = "---"

wb.save(filename=dest_filename)

df = pd.read_excel(user + '\\Downloads\\Teste.xlsx', engine="openpyxl")

acumulado_df = pd.pivot_table(df, values=['Qtd. Viagens'],index=['Macro Área Origem','Macro Área Destino','Área Origem','Área Destino'],aggfunc='sum')

writer = pd.ExcelWriter(user + "\\Downloads\\Produto Final.xlsx", engine='xlsxwriter')

acumulado_df.to_excel(writer, sheet_name='Final',merge_cells=False)

writer.save()

os.remove(user + '\\Downloads\\Teste.xlsx')

