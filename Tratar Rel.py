from openpyxl import load_workbook
from openpyxl.styles import Alignment, Side, Border, Font, PatternFill
import pyperclip, pyautogui, os, datetime
from suporte.Recursos import TelaMesclar
from suporte.Recursos import Loop
from time import sleep
import pandas as pd
from datetime import date
import xlsxwriter

user = os.path.expanduser('~')

tela = TelaMesclar()
tela.Iniciar()

center_align = Alignment(horizontal='center', vertical='center')
low_align = Alignment(horizontal='center', vertical='bottom')
box = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
box2 = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))

verde = PatternFill('solid', fgColor='6be07f')
azul = PatternFill('solid', fgColor='8589f0')
vermelho = PatternFill('solid', fgColor='e37373')
amarelo = PatternFill('solid', fgColor='efee9b')
laranja = PatternFill('solid', fgColor='e3b973')
roxo = PatternFill('solid', fgColor='9e58c7')
rosa = PatternFill('solid', fgColor='ff6fff')
transparente = PatternFill('none')

if tela.values['relatorio_resumo'] == True and tela.values['relatorio_ocupacao'] == False and tela.values['relatorio_carretinhas'] == False and tela.values['relatorio_permanencia'] == False: 
    
    wb = load_workbook(filename=(user + '\\Downloads\\RelatorioDisponibilidadeCompleto.xlsx'))
    wb2 = load_workbook(filename=(user + '\\Downloads\\RelatorioDisponibilidadeCompleto (1).xlsx'))
    
    dest_filename = user + '\\Downloads\\Barreiro e Jeceaba (Relatório Disponibilidade).xlsx'
    
    ws = wb['Sheet1']
    ws2 = wb2['Sheet1']
            
    max_row = ws.max_row
    max_col = ws.max_column
    max_row_ws2 = ws2.max_row
    
    Loop.Duplo(max_row_ws2, max_col, max_row, ws2, ws, center_align, box)
    max_row = ws.max_row

    ws.delete_cols(9)
    max_col = ws.max_column

    for i in range(1, max_row + 1):
        if ws.cell(i, 1).value == 'Data':
            aux = i + 1
            for n in range(aux, max_row + 1):
                ws.cell(n, 1).value = datetime.datetime.strptime(ws.cell(n, 1).value, '%d/%m/%Y').date()
            break
    
    wb.save(filename=dest_filename)
    
    pyperclip.copy(user + '\\Downloads\\Barreiro e Jeceaba (Relatório Disponibilidade).xlsx')
    sleep(1)
    pyautogui.hotkey('win', 'r')
    sleep(1)
    pyautogui.hotkey('ctrl', 'v')
    sleep(1)
    pyautogui.press('enter')

    # os.remove(user + '\\Downloads\\RelatorioDisponibilidadeCompleto.xlsx')
    # os.remove(user + '\\Downloads\\RelatorioDisponibilidadeCompleto (1).xlsx')

elif tela.values['relatorio_resumo'] == False and tela.values['relatorio_ocupacao'] == True and tela.values['relatorio_carretinhas'] == False and tela.values['relatorio_permanencia'] == False:
    
    wb = load_workbook(filename=(user + '\\Downloads\\Relatorio_produtividade_sessao.xlsx'))
    wb2 = load_workbook(filename=(user + '\\Downloads\\Relatorio_produtividade_sessao (1).xlsx'))
    wb3 = load_workbook(filename=(user + '\\Downloads\\Relatorio_produtividade_sessao (2).xlsx'))
    wb4 = load_workbook(filename=(user + '\\Downloads\\Relatorio_produtividade_sessao (3).xlsx'))
    
    dest_filename = user + '\\Downloads\\Produtividade_Sessao (BAR + JCB).xlsx'
    
    ws = wb['Sheet1']
    ws2 = wb2['Sheet1']
    ws3 = wb3['Sheet1']
    ws4 = wb4['Sheet1']
    
    format = "%d/%m/%Y %H:%M:%S"

    max_row2 = ws2.max_row
    max_col2 = ws2.max_column
    max_row_ws = ws.max_row
    max_row_ws3 = ws3.max_row
    max_row_ws4 = ws4.max_row
    
    Loop.Duplo2(5, max_row2, 1, max_col2, max_row_ws3, ws2, ws3, center_align, box)
    max_row2 = ws2.max_row
    
    Loop.Duplo2(5, max_row2, 1, max_col2, max_row_ws4, ws2, ws4, center_align, box)
    max_row2 = ws2.max_row

    ws2.delete_cols(14, 9)
    ws2.insert_cols(2, 1)
    
    Loop.Unico(2, 5, max_row2, ws2, center_align, box, 'Barreiro-Tradimaq')
    
    ws2.cell(4, 2).value = 'Identificador'
    ws2.cell(4, 2).alignment = low_align
    ws2.cell(4, 2).border = box2
    ws2.cell(4, 2).font = Font(bold=True)
    
    max_col2 = ws2.max_column
    
    Loop.Condicional(6, 5, max_row2, ws2, center_align, box, 'CAMINHÃO BÁSCULA', 'BASCULA', amarelo)
    
    Loop.CondicionalDuplo(4, 6, 5, max_row2, ws2, center_align, box, 'AJUSTAGEM BARRAS', 'PLATSUL', 'RUCKER', verde)
    Loop.CondicionalDuplo(4, 6, 5, max_row2, ws2, center_align, box, 'AJUSTAGEM BARRAS', 'PLATSUL', 'EMPILHADEIRA', verde)

    Loop.CondicionalDuploInverso(4, 3, 5, max_row2, ws2, center_align, box, 'Turno II', 'POOL', 'Turno II', azul)
    Loop.CondicionalDuploInverso(4, 3, 5, max_row2, ws2, center_align, box, 'Turno III', 'OFICINA DE PONTE ROLANTE', 'Turno III', azul)
    
    Loop.Condicional2(4, 6, 5, max_row2, ws2, center_align, box, 'PORTICO TRIAGEM', 'PÁ CARREGADEIRA', transparente)
    
    Loop.Condicional3(4, 6, 6, 5, max_row2, ws2, center_align, box, 'EMPILHADEIRA - 4,5T', 'LUVAS', 'EMPILHADEIRA')
    Loop.Condicional3(4, 6, 6, 5, max_row2, ws2, center_align, box, 'EMPILHADEIRA - 2,5T', 'POOL', 'EMPILHADEIRA')
    Loop.Condicional3(4, 6, 6, 5, max_row2, ws2, center_align, box, 'EMPILHADEIRA - 4,5T', 'GRANDE PARADA', 'EMPILHADEIRA')
    Loop.Condicional3(4, 6, 6, 5, max_row2, ws2, center_align, box, 'EMPILHADEIRA - 2,5T', 'LABORATÓRIOPD', 'EMPILHADEIRA')
    Loop.Condicional3(4, 6, 6, 5, max_row2, ws2, center_align, box, 'EMPILHADEIRA - 2,5T', 'LOCAL 34', 'EMPILHADEIRA')
    Loop.Condicional3(4, 6, 6, 5, max_row2, ws2, center_align, box, 'EMPILHADEIRA - 4,5T', 'LOCAL40', 'EMPILHADEIRA')
    Loop.Condicional3(4, 6, 5, 5, max_row2, ws2, center_align, box, 'EMPILHADEIRA - 2,5T', 'OFCENTRAL', 'E2')
    Loop.Condicional3(4, 6, 5, 5, max_row2, ws2, center_align, box, 'EMPILHADEIRA - 4,5T', 'OFCENTRAL', 'E4')
    Loop.Condicional3(4, 6, 5, 5, max_row2, ws2, center_align, box, 'EMPILHADEIRA - 2,5T', 'ALMOXARIFADO CENTRAL', 'E2')
    Loop.Condicional3(4, 6, 5, 5, max_row2, ws2, center_align, box, 'EMPILHADEIRA - 4,5T', 'ALMOXARIFADO CENTRAL', 'E4')
    Loop.Condicional3(4, 6, 5, 5, max_row2, ws2, center_align, box, 'EMPILHADEIRA - 4,5T', 'ALMOXARIFADO CENTRAL', 'E5')
    Loop.Condicional3(4, 6, 5, 5, max_row2, ws2, center_align, box, 'EMPILHADEIRA - 2,5T', 'ALMOXARIFADO CENTRAL', 'E3')
    Loop.Condicional3(4, 6, 5, 5, max_row2, ws2, center_align, box, 'EMPILHADEIRA - 7T', 'OFCENTRAL', 'E7')
    Loop.Condicional3(4, 6, 5, 5, max_row2, ws2, center_align, box, 'EMPILHADEIRA - 2,5T', 'PATIOCRL', 'E2')
    Loop.Condicional3(4, 6, 5, 5, max_row2, ws2, center_align, box, 'EMPILHADEIRA - 7T', 'PATIOCRL', 'E7')
    Loop.Condicional3(4, 6, 5, 5, max_row2, ws2, center_align, box, 'EMPILHADEIRA - 7T', 'PATIOCRL', 'E9')
    Loop.Condicional3(4, 6, 6, 5, max_row2, ws2, center_align, box, 'EMPILHADEIRA - 7T', 'BREDERO', 'EMPILHADEIRA')
    Loop.Condicional3(4, 6, 6, 5, max_row2, ws2, center_align, box, 'EMPILHADEIRA - 7T', 'AJUSTAGEM BARRAS', 'EMPILHADEIRA')
    Loop.Condicional3(4, 6, 6, 5, max_row2, ws2, center_align, box, 'EMPILHADEIRA - 7T', 'PATIOST1', 'EMPILHADEIRA')
    Loop.Condicional3(4, 6, 6, 5, max_row2, ws2, center_align, box, 'EMPILHADEIRA - 7T', 'FORJA MATÉRIA PRIMA', 'EMPILHADEIRA')
    Loop.Condicional3(4, 6, 6, 5, max_row2, ws2, center_align, box, 'RETROESCAVADEIRA', 'ACIARIA SUCATA', 'PÁ CARREGADEIRA')
    
    Loop.Duplo2(5, max_row2, 1, max_col2, max_row_ws, ws2, ws, center_align, box)
    max_row2 = ws2.max_row
    
    Loop.Condicional(3, 5, max_row2, ws2, center_align, box, 'Turno I', 'Turno 1', transparente)
    Loop.Condicional(3, 5, max_row2, ws2, center_align, box, 'Turno II', 'Turno 2', transparente)
    Loop.Condicional(3, 5, max_row2, ws2, center_align, box, 'Turno III', 'Turno 3', transparente)
    
    Loop.Condicional3(2, 6, 6, 5, max_row2, ws2, center_align, box, 'MINI-RETRO', 'Jeceaba-Tradimaq', 'RETRO ESCAVADEIRA')
    Loop.Condicional3(2, 6, 5, 5, max_row2, ws2, center_align, box, 'CAMINHÃO ROLL-ON', 'Barreiro-Tradimaq', 'CR')
    Loop.Condicional3(2, 6, 5, 5, max_row2, ws2, center_align, box, 'CAMINHÃO BÁSCULA', 'Jeceaba-Tradimaq', 'CBA')
    Loop.Condicional4(2, 6, 6, 5, max_row2, ws2, center_align, box, 'MINI-RETRO', 'Jeceaba-Tradimaq', 'CARREGADEIRA')

    Loop.Deletar(5, 5, max_row2, ws2, 'SPIN')
    max_row2 = ws2.max_row

    Loop.Deletar(4, 5, max_row2, ws2, 'TRADIMAQ')
    max_row2 = ws2.max_row

    Loop.Deletar(4, 5, max_row2, ws2, 'TRADIMAQ TREINAMENTO')
    max_row2 = ws2.max_row
    
    Loop.Deletar(5, 5, max_row2, ws2, 'SPOT')
    max_row2 = ws2.max_row

    Loop.Deletar(5, 5, max_row2, ws2, 'Strada')
    max_row2 = ws2.max_row

    Loop.Deletar(6, 5, max_row2, ws2, 'COMBOIO')
    max_row2 = ws2.max_row

    Loop.Deletar(6, 5, max_row2, ws2, 'CARRO')
    max_row2 = ws2.max_row

    Loop.Deletar(5, 5, max_row2, ws2, 'TCM2921')
    max_row2 = ws2.max_row
    
    Loop.Deletar(8, 5, max_row2, ws2, 'DDMX')
    max_row2 = ws2.max_row  
    
    Loop.Deletar2(8, 5, max_row2, ws2, 'TESTE')
    max_row2 = ws2.max_row

    Loop.Deletar2(5, 5, max_row2, ws2, 'RUB8D88 - SPIN - TRADIMAQ')
    max_row2 = ws2.max_row

    Loop.Deletar2(5, 5, max_row2, ws2, 'V. ESP - TP7-20')
    max_row2 = ws2.max_row

    Loop.Deletar2(5, 5, max_row2, ws2, 'P7-12')
    max_row2 = ws2.max_row

    Loop.Deletar2(5, 5, max_row2, ws2, 'P5-3')
    max_row2 = ws2.max_row
    
    Loop.Deletar2(8, 5, max_row2, ws2, 'Paulo Queiroz')
    max_row2 = ws2.max_row

    Loop.Deletar2(8, 5, max_row2, ws2, 'Luiz Francisco Barbosa')
    max_row2 = ws2.max_row

    Loop.Deletar2(4, 5, max_row2, ws2, 'EQUIPE DESCARGA')
    max_row2 = ws2.max_row
    
    Loop.Deletar2(5, 5, max_row2, ws2, '6425723 (TROCADO)')
    max_row2 = ws2.max_row

    i = 5
    while i <= max_row2:
        celula = ws2.cell(row=i, column=12)
        celula2 = ws2.cell(row=i, column=13)
        if celula2.value is not None and celula.value is not None:
            if celula2.value > celula.value:
                celula2.value = celula.value
        i += 1

    ws2_2 = wb2.create_sheet('Sheet_A')
    ws2_2.title = 'Sheet2'
    ws2_3 = wb2.create_sheet('Sheet_B')
    ws2_3.title = 'Sheet3'
    
    for i in range(1, 14):
        ws2_2.cell(1, i).value = ws2.cell(4, i).value
        ws2_2.cell(1, i).alignment = low_align
        ws2_2.cell(1, i).border = box2
        ws2_2.cell(1, i).font = Font(bold=True)

    for i in range(1, 6):
        ws2_3.cell(1, i).value = ws2.cell(4, i).value
        ws2_3.cell(1, i).alignment = low_align
        ws2_3.cell(1, i).border = box2
        ws2_3.cell(1, i).font = Font(bold=True)

    max_row2_2 = ws2_2.max_row
    max_row2_3 = ws2_3.max_row
    
    i = 5
    n = 2
    
    while i <= max_row2:
        if ws2.cell(i, 13).value == datetime.time(0, 0, 0):
            for j in range(1, 14):
                ws2_2.cell(n, j).value = ws2.cell(i, j).value
                ws2_2.cell(n, j).alignment = center_align
                ws2_2.cell(n, j).border = box
            n += 1
        i += 1

    for j in range(1, 6):
        ws2_3.cell(2, j).value = ws2_2.cell(2, j).value
        ws2_3.cell(2, j).alignment = center_align
        ws2_3.cell(2, j).border = box

    max_row2_2 = ws2_2.max_row
    max_row2_3 = ws2_3.max_row
    
    i = 2
    
    while i <= max_row2_2:
        aux = 2
        for n in range(2, max_row2_3 + 1):
            if ws2_3.cell(n, 5).value != ws2_2.cell(i, 5).value:
                aux += 1
                if aux == max_row2_3 + 1:
                    for j in range(1, 6):
                        ws2_3.cell(max_row2_3 + 1, j).value = ws2_2.cell(i, j).value
                        ws2_3.cell(max_row2_3 + 1, j).alignment = center_align
                        ws2_3.cell(max_row2_3 + 1, j).border = box
                    max_row2_3 = ws2_3.max_row
        i+=1      

    max_row2_3 = ws2_3.max_row
    
    ws2_3.cell(max_row2_3 + 2, 1).value = 'POSSÍVEIS EQUIPAMENTOS COM DEFEITO'
    ws2_3.cell(max_row2_3 + 2, 1).alignment = low_align
    ws2_3.cell(max_row2_3 + 2, 1).border = box2
    ws2_3.cell(max_row2_3 + 2, 1).font = Font(bold=True)
    
    max_row2_3 = ws2_3.max_row
    max_row2 = ws2.max_row
    
    for n in range(2, max_row2_3):
        aux = 0
        aux2 = 0
        for i in range(5, max_row2):
            if ws2.cell(i, 5).value == ws2_3.cell(n, 5).value:
                aux += 1

        for i in range(5, max_row2):
            if ws2.cell(i, 5).value == ws2_3.cell(n, 5).value and ws2.cell(i, 13).value == datetime.time(0, 0, 0):
                aux2 += 1

        if aux == aux2:
            ws2_3.cell(max_row2_3 + 1, 1).value = ws2_3.cell(n, 5).value
            ws2_3.cell(max_row2_3 + 1, 1).alignment = center_align
            ws2_3.cell(max_row2_3 + 1, 1).border = box
            max_row2_3 = ws2_3.max_row

    Loop.Deletar4(13, 5, max_row2, ws2, datetime.time(0, 30, 0))
    max_row2 = ws2.max_row
    
    Loop.Deletar4(12, 5, max_row2, ws2, datetime.time(0, 30, 0))
    max_row2 = ws2.max_row

    Loop.Deletar4(11, 5, max_row2, ws2, datetime.time(0, 30, 0))
    max_row2 = ws2.max_row

    for i in range(5, max_row2 + 1):
        if ws2.cell(i, 1).value == None:
            pass
        else:
            ws2.cell(i, 1).value = datetime.datetime.strptime(ws2.cell(i, 1).value, '%d/%m/%Y').date()

    wb2.save(filename=dest_filename)
    sleep(5)

    # os.remove(user + '\\Downloads\\Relatorio_produtividade_sessao.xlsx')
    # os.remove(user + '\\Downloads\\Relatorio_produtividade_sessao (1).xlsx')
    # os.remove(user + '\\Downloads\\Relatorio_produtividade_sessao (2).xlsx')
    # os.remove(user + '\\Downloads\\Relatorio_produtividade_sessao (3).xlsx')

    pyperclip.copy(user + '\\Downloads\\Produtividade_Sessao (BAR + JCB).xlsx')
    sleep(1)
    pyautogui.hotkey('win', 'r')
    sleep(1)
    pyautogui.hotkey('ctrl', 'v')
    sleep(1)
    pyautogui.press('enter')

elif tela.values['relatorio_resumo'] == False and tela.values['relatorio_ocupacao'] == False and tela.values['relatorio_carretinhas'] == True and tela.values['relatorio_permanencia'] == False:

    center_align = Alignment(horizontal='center', vertical='center')
    box = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    wb = load_workbook(filename=(user + '\\Downloads\\Relatorio_viagens_carretinhas_ruckers.xlsx'))

    ws = wb['Sheet1']

    ws.delete_cols(3, 3)
    ws.delete_cols(4, 10)
    ws.delete_cols(5, 2)
    ws.delete_cols(7, 5)
    ws.delete_cols(8, 1)
    ws.delete_cols(9, 9)
    ws.delete_cols(10, 5)

    ws.cell(4, 10).value = 'Viagens'
    ws.cell(4, 10).alignment = low_align
    ws.cell(4, 10).border = box2
    ws.cell(4, 10).font = Font(bold=True)

    ws.cell(4, 11).value = 'Meta'
    ws.cell(4, 11).alignment = low_align
    ws.cell(4, 11).border = box2
    ws.cell(4, 11).font = Font(bold=True)

    max_row_ws = ws.max_row

    Loop.Unico(10, 5, max_row_ws, ws, center_align, box, 1)

    Loop.Unico(11, 5, max_row_ws, ws, center_align, box, 1)

    max_row_ws = ws.max_row
    
    for i in range(5, max_row_ws + 1):
        ws.cell(i, 1).value = datetime.datetime.strptime(ws.cell(i, 1).value, '%d/%m/%Y').date()

    wb.save(filename=(user + '\\Downloads\\Viagens Carretinhas.xlsx'))

    os.remove(user + '\\Downloads\\Relatorio_viagens_carretinhas_ruckers.xlsx')

#    pyperclip.copy(user + '\\Downloads\\Viagens Carretinhas.xlsx')
#    pyautogui.hotkey('win', 'r')
#    sleep(1)
#    pyautogui.hotkey('ctrl', 'v')
#    sleep(1)
#    pyautogui.press('enter')
    
#    pyperclip.copy('\\\\srvoffice\\PL\\PLM\\Coord_Movimentação_Equipamentos\\10. Power BI\\Logística Interna - Indicadores\\Acompanhamento Diário')
#    pyautogui.hotkey('win', 'r')
#    sleep(1)
#    pyautogui.hotkey('ctrl', 'v')
#    sleep(1)
#    pyautogui.press('enter')

elif tela.values['relatorio_resumo'] == False and tela.values['relatorio_ocupacao'] == False and tela.values['relatorio_carretinhas'] == False and tela.values['relatorio_permanencia'] == True:
    wb = load_workbook(filename = user + r'\Downloads\Relatorio_permanencia_carreta_tempo.xlsx')
    wb2 = load_workbook(filename = user + r'\Downloads\Relatorio_permanencia_carreta_tempo (1).xlsx')
    wb3 = load_workbook(r'\\srvoffice\PL\PLM\Coord_Movimentação_Equipamentos\10. Power BI\Logística Interna - Indicadores\Acompanhamento Diário\Relatorio_permanencia_carreta_tempo.xlsx')

    ws = wb['Sheet1']
    ws2 = wb2['Sheet1']
    ws3 = wb3['Sheet1']

    max_row2 = ws2.max_row
    max_row3 = ws3.max_row
    max_row_ws = ws.max_row
    max_col3 = ws3.max_column

    Loop.Unico(4, 5, max_row2, ws2, center_align, box, 'PLATSUL')

    Loop.Unico(4, 5, max_row_ws, ws, center_align, box, 'DPA')
        
    max_col2 = ws2.max_column
    max_col = ws.max_column
        
    Loop.Duplo2(5, max_row2, 1, max_col2, max_row_ws, ws2, ws, center_align, box)
    max_row2 = ws2.max_row

    Loop.Duplo2(5, max_row3, 1, max_col3, max_row2, ws3, ws2, center_align, box)
    max_row3 = ws3.max_row

    wb3.save(filename = r'\\srvoffice\PL\PLM\Coord_Movimentação_Equipamentos\10. Power BI\Logística Interna - Indicadores\Acompanhamento Diário\Relatorio_permanencia_carreta_tempo.xlsx')

    pyperclip.copy(r'\\srvoffice\PL\PLM\Coord_Movimentação_Equipamentos\10. Power BI\Logística Interna - Indicadores\Acompanhamento Diário\Carretas.pbix')
    pyautogui.hotkey("win","r")
    sleep(1)
    pyautogui.hotkey("ctrl","v")
    sleep(1)
    pyautogui.press("enter")

    os.remove(user + r'\Downloads\Relatorio_permanencia_carreta_tempo (1).xlsx')
    os.remove(user + r'\Downloads\Relatorio_permanencia_carreta_tempo.xlsx')

elif tela.values['relatorio_manutencao'] == True:
    
    center_align = Alignment(horizontal='center', vertical='center')
    box = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    data_hoje = date.today().strftime("%d/%m/%Y")

    user = os.path.expanduser('~')

    barreiro_df = pd.read_excel(user + r'\Downloads\barreiro.xls')
    jeceaba_df = pd.read_excel(user + r'\Downloads\jeceaba.xls')

    linha = barreiro_df[barreiro_df['Unnamed: 0'] == 'Veículos dentro das Áreas'].index[0]

    barreiro_df.drop(barreiro_df.columns[10:17], axis=1, inplace=True)

    barreiro_df.drop(barreiro_df.index[:linha+1], axis=0, inplace=True)

    barreiro_df.columns = barreiro_df.iloc[0]

    barreiro_df = barreiro_df[1:]

    barreiro_df = barreiro_df.rename(columns={'Sub-grupo': 'Data', 'Grupo da Área': 'Identificador'})

    barreiro_df['Data'] = data_hoje

    barreiro_df['Identificador'] = 'Barreiro-Tradimaq'

    barreiro_df = barreiro_df[barreiro_df['Categoria'].str.contains('OFICINA', na=False)]

    linha = jeceaba_df[jeceaba_df['Unnamed: 0'] == 'Veículos dentro das Áreas'].index[0]

    jeceaba_df.drop(jeceaba_df.columns[10:17], axis=1, inplace=True)

    jeceaba_df.drop(jeceaba_df.index[:linha+1], axis=0, inplace=True)

    jeceaba_df.columns = jeceaba_df.iloc[0]

    jeceaba_df = jeceaba_df[1:]

    jeceaba_df = jeceaba_df.rename(columns={'Sub-grupo': 'Data', 'Grupo da Área': 'Identificador'})

    jeceaba_df['Data'] = data_hoje

    jeceaba_df['Identificador'] = 'Jeceaba-Tradimaq'

    frames = [barreiro_df, jeceaba_df]

    result = pd.concat(frames)

    result = result.reset_index(drop=True)

    writer = pd.ExcelWriter(user + r'\Downloads\Equipamentos em Manutenção.xlsx', engine='xlsxwriter')
    result.to_excel(writer, sheet_name='Sheet1')
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']

    border_fmt = workbook.add_format({'bottom':1, 'top':1, 'left':1, 'right':1})
    worksheet.conditional_format(xlsxwriter.utility.xl_range(0, 0, len(result), len(result.columns)), {'type': 'no_errors', 'format': border_fmt})
    writer.save()

    wb = load_workbook(filename=(user + '\\Downloads\\Equipamentos em Manutenção.xlsx'))

    ws = wb['Sheet1']

    ws.delete_cols(1, 1)

    for i in range(2, ws.max_row + 1):
        ws.cell(i, 2).value = datetime.datetime.strptime(ws.cell(i, 2).value, '%d/%m/%Y').date()

    os.remove(user + '\\Downloads\\barreiro.xls')
    os.remove(user + '\\Downloads\\jeceaba.xls')

    wb.save(filename=(user + '\\Downloads\\Equipamentos em Manutenção.xlsx'))

elif tela.values['relatorio_ausencias'] == True:

    wb = load_workbook(filename = user + r'\Downloads\Relatorio_ausencias.xlsx')
    wb2 = load_workbook(filename = user + r'\Downloads\Relatorio_ausencias (1).xlsx')

    ws = wb['Sheet1']
    ws2 = wb2['Sheet1']

    max_row = ws.max_row
    max_row2 = ws2.max_row
    max_col = ws.max_column
    max_col2 = ws2.max_column

    Loop.Duplo2(5, max_row, 1, max_col, max_row2, ws, ws2, center_align, box)
    max_row = ws.max_row

    for i in range(5, ws.max_row + 1):
        ws.cell(i, 1).value = datetime.datetime.strptime(ws.cell(i, 1).value, '%d/%m/%Y').date()

    for i in range(5, ws.max_row + 1):
        ws.cell(i, 10).value = datetime.datetime.strptime(ws.cell(i, 11).value, '%d/%m/%Y %H:%M:%S')

    for i in range(5, ws.max_row + 1):
        ws.cell(i, 11).value = datetime.datetime.strptime(ws.cell(i, 11).value, '%d/%m/%Y %H:%M:%S')

    for i in range(5, ws.max_row + 1):
        ws.cell(i, 12).value = datetime.datetime.strptime(ws.cell(i, 12).value, '%H:%M:%S').time()

    wb.save(filename=(user + '\\Downloads\\Ausências.xlsx'))

    os.remove(user + '\\Downloads\\Relatorio_ausencias.xlsx')
    os.remove(user + '\\Downloads\\Relatorio_ausencias (1).xlsx')

    pyperclip.copy(user + '\\Downloads\\Ausências.xlsx')
    sleep(1)
    pyautogui.hotkey('win', 'r')
    sleep(1)
    pyautogui.hotkey('ctrl', 'v')
    sleep(1)
    pyautogui.press('enter')
else:
    pass