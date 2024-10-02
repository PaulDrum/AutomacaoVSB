import PySimpleGUI as sg, os
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

user = os.path.expanduser('~')

class TelaPython:

    def __init__(self):
        usuarios = [
         'VITOR.VALADARES', 'lucas.saraujo']
        senhas = ['Vitor@2022', 'Lucas@2022']
        user_win = ['C:\\Users\\Vitor.Valadares', 'C:\\Users\\s-Lucas.SAraujo']
        user = os.path.expanduser('~')
        usuario = ''
        senha = ''
        for i in range(0, 2):
            if user == user_win[i]:
                usuario = usuarios[i]
                senha = senhas[i]

        sg.theme('Reddit')
        layout = [
         [
          sg.Text('Usuário:', size=(12, 0)), sg.Input(default_text=usuario, key='usuario', size=(22, 0))],
         [
          sg.Text('Senha:', size=(12, 0)), sg.Input(default_text=senha, key='senha', password_char='*', size=(22, 0))],
         [
          sg.Text('Data Inicial:', size=(12, 0)), sg.Input(key='data_inicial', size=(22, 0))],
         [
          sg.Text('Data Final:', size=(12, 0)), sg.Input(key='data_final', size=(22, 0))],
         [
          sg.Text('Tipo de Relatório:')],
         [
          sg.Radio('Disponibilidade Contratual', 'relatórios', key='relatorio_resumo'), sg.Radio('Ocupação', 'relatórios', key='relatorio_ocupacao')],
         [
          sg.Radio('Viagens Carretinhas', 'relatórios', key='relatorio_carretinhas'), sg.Radio('Permanência', 'relatórios', key='relatorio_permanencia')],
                   [
          sg.Radio('Equipamentos em Manutenção', 'relatórios', key='relatorio_manutencao')],
          [
          sg.Radio('Ausências', 'relatórios', key='relatorio_ausencias')],
         [
          sg.Button('Enviar Dados')]]
        self.janela = sg.Window('Geração de Relatório').layout(layout)
        self.button, self.values = self.janela.Read()

    def Iniciar(self):
        self.janela.close()

class AcesseMeuSite:

    def Logar(navegador, conta, usuario, senha):
        navegador.find_element(By.ID, "login-conta").send_keys(conta)
        navegador.find_element(By.ID, "login-email").send_keys(usuario)
        navegador.find_element(By.ID, "login-senha").send_keys(senha)
        navegador.find_element(By.ID, 'button-login').click()

class AcesseMeuSiteMonitor:

    def Logar(navegador, conta, usuario, senha):
        navegador.find_element(By.ID, "mat-input-0").send_keys(conta)
        navegador.find_element(By.ID, "mat-input-1").send_keys(usuario)
        navegador.find_element(By.ID, "mat-input-2").send_keys(senha)
        navegador.find_element(By.XPATH, "//*[contains(text(),'Acessar')]").click()

class AcesseMeuSiteTriyax:

    def Logar(navegador, conta, usuario, senha):
        navegador.find_element(By.ID, "mat-input-0").send_keys(conta)
        navegador.find_element(By.ID, "mat-input-1").send_keys(usuario)
        navegador.find_element(By.ID, "mat-input-2").send_keys(senha)
        navegador.find_element(By.XPATH, "//button").click()

class AcesseMeuSiteAnalitico:

    def Logar(navegador, usuario, senha):
        navegador.find_element(By.XPATH,'//input[@name="login"]').send_keys(usuario)
        navegador.find_element(By.XPATH,'//input[@name="senha"]').send_keys(senha)
        navegador.find_element(By.XPATH,'//button[@type="submit"]').click()

class TelaMesclar:

    def __init__(self):
        sg.theme('Reddit')
        layout = [
         [
            sg.Text('Tipo de Relatório:')],
         [
            sg.Radio('Disponibilidade Contratual', 'relatórios', key='relatorio_resumo'), sg.Radio('Ocupação', 'relatórios', key='relatorio_ocupacao')],
         [
            sg.Radio('Viagens Carretinhas', 'relatórios', key='relatorio_carretinhas'), sg.Radio('Permanência', 'relatórios', key='relatorio_permanencia')],
            [
            sg.Radio('Equipamentos em Manutenção', 'relatórios', key='relatorio_manutencao')],
            [
            sg.Radio('Ausências', 'relatórios', key='relatorio_ausencias')],
         [
          sg.Button('Mesclar')]]
        self.janela = sg.Window('Mesclar Planilhas').layout(layout)
        self.button, self.values = self.janela.Read()

    def Iniciar(self):
        self.janela.close()

class TelaJornadas:

    def __init__(self):
        sg.theme('Reddit')
        layout = [
            [sg.Text('----- INFORMAÇÕES DE ACESSO -----', auto_size_text=True)],
            [sg.Text('Usuário:', size=(17, 0)), sg.Input(key='usuario', size=(20, 0))],
            [sg.Text('Senha:', size=(17, 0)), sg.Input(key='senha', password_char='*', size=(20, 0))],
            [sg.Text('----- DATA DA PROGRAMAÇÃO -----', auto_size_text=True)],
            [sg.Text("Data Inicial:",size=(17,0)), sg.In(key="-CAL-", enable_events=True, font=('Arial', 11), size=(17,0)), sg.CalendarButton('Data', key='calendario', format=('%d/%m/%Y'), size=(6,0), font=('Arial', 10))],
            [sg.Text("Data Final:",size=(17,0)), sg.In(key="-CAL2-", enable_events=True, font=('Arial', 11), size=(17,0)), sg.CalendarButton('Data', key='calendario2', format=('%d/%m/%Y'), size=(6,0), font=('Arial', 10))],
            [sg.Text('----- HORÁRIO DOS TURNOS -----', auto_size_text=True)],
            [sg.Text('Turno 1 Fim', size=(17, 0)), sg.Input(key='t1fim', size=(20, 0))],
            [sg.Text('Turno 2 Inicio:', size=(17, 0)), sg.Input(key='t2inicio', size=(20, 0))],
            [sg.Text('Turno 2 Fim:', size=(17, 0)), sg.Input(key='t2fim', size=(20, 0))],
            [sg.Text('Turno 3 Inicio:', size=(17, 0)), sg.Input(key='t3inicio', size=(20, 0))],
            [sg.Text('12x36 Dia Inicio:', size=(17, 0)), sg.Input(key='12x36dinicio', size=(20, 0))],
            [sg.Text('12x36 Dia Fim:', size=(17, 0)), sg.Input(key='12x36dfim', size=(20, 0))],
            [sg.Text('12x36 Noite Inicio:', size=(17, 0)), sg.Input(key='12x36ninicio', size=(20, 0))],
            [sg.Text('12x36 Noite Fim:', size=(17, 0)), sg.Input(key='12x36nfim', size=(20, 0))],
            [sg.Text('----- ESCALAS -----', auto_size_text=True)],
            [sg.CB('6x1 e 6x2', font=('Arial', 12),key='6x1'), sg.CB('12x36', font=('Arial', 12),key='12x36')],
            [sg.Text('----- BASE DE DADOS -----', auto_size_text=True)],
            [sg.Text('Programação de Jornada.xlsx:', size=(22,0)), sg.In(size=(20,1) ,key='doc'), sg.FileBrowse(font=('Arial', 12))],
            [sg.Button('Programar', font=('Arial', 12)), sg.Quit(font=('Arial', 12))]]
        
        
        self.janela = sg.Window('Programação de Jornada', layout, resizable = True).Finalize()

        while True:

                self.button, self.values = self.janela.read()

                if self.button in (sg.WINDOW_CLOSED, "Quit"):
                    break
                elif self.button == "Programar":
                    if self.values["-CAL-"] and self.values["-CAL2-"] != '':
                         if self.values["usuario"] and self.values["senha"] != '':
                              if self.values["6x1"] or self.values["12x36"] != False:
                                    if self.values["doc"] != '':
                                        break
                                    else:
                                        sg.popup_error('Escolha uma base de dados!')
                              else:
                                sg.popup_error('Escolha uma escala!')
                         else:
                            sg.popup_error('Preencha as informações de acesso!')
                    else:
                         sg.popup_error('Preencha as datas!')

class TelaCarretas:

    def __init__(self):
        sg.theme('Reddit')
        layout = [
            [sg.Text('----- INFORMAÇÕES DE ACESSO DA MÁQUINA-----', auto_size_text=True)],
            [sg.Text('E-Mail:', size=(17, 0)), sg.Input(key='usuario', size=(20, 0))],
            [sg.Text('Senha:', size=(17, 0)), sg.Input(key='senha', size=(20, 0))],
            [sg.Text('----- DATA NECESSÁRIA DOS DADOS -----', auto_size_text=True)],
            [sg.Text("Data Inicial:",size=(17,0)), sg.In(key="-CAL-", enable_events=True, font=('Arial', 11), size=(17,0)), sg.CalendarButton('Data', key='calendario', format=('%d/%m/%Y'), size=(6,0), font=('Arial', 10))],
            [sg.Text("Data Final:",size=(17,0)), sg.In(key="-CAL2-", enable_events=True, font=('Arial', 11), size=(17,0)), sg.CalendarButton('Data', key='calendario2', format=('%d/%m/%Y'), size=(6,0), font=('Arial', 10))],
            [sg.Button('Atualizar', font=('Arial', 12)), sg.Quit(font=('Arial', 12))]]
        
        self.janela = sg.Window('Atualização Dados Carreta', layout, resizable = True).Finalize()

        while True:

                self.button, self.values = self.janela.read()

                if self.button in (sg.WINDOW_CLOSED, "Quit"):
                    break
                elif self.button == "Atualizar":
                    if self.values["-CAL-"] and self.values["-CAL2-"] != '':
                        if self.values["usuario"] and self.values["senha"] != '':
                            break
                        else:
                            sg.popup_error('Preencha as informações de acesso!')
                    else:
                         sg.popup_error('Preencha as datas!')

    def Iniciar(self):
        self.janela.close()

class TelaCarretasso:

    def __init__(self):
        sg.theme('Reddit')
        layout = [
            [sg.Text('----- BASE DE DADOS -----', auto_size_text=True)],
            [sg.Text('Viagens de carretinhas.xlsx:', size=(22,0)), sg.In(size=(20,1) ,key='doc'), sg.FileBrowse(font=('Arial', 12))],
            [sg.Button('Gerar Dados', font=('Arial', 12)), sg.Quit(font=('Arial', 12))]]
        
        self.janela = sg.Window('Atualização Dados Carreta', layout, resizable = True).Finalize()

        while True:

                self.button, self.values = self.janela.read()

                if self.button in (sg.WINDOW_CLOSED, "Quit"):
                    break
                elif self.button == "Gerar Dados":
                    if self.values["doc"] != '':
                        break
                    else:
                        sg.popup_error('Informe a base de dados!')

    def Iniciar(self):
        self.janela.close()

class TelaMetrica:

    def __init__(self):
        sg.theme('Reddit')
        layout = [
            [sg.Text('Disponibilidade Contratual:', size=(22,0), font=('Arial', 12)), sg.In(size=(20,1) ,key='disponibilidade', font=('Arial', 12)), sg.FileBrowse(font=('Arial', 12))],
            [sg.Text('Relatório de Ausências:', size=(22,0), font=('Arial', 12)), sg.In(size=(20,1) ,key='ausencia', font=('Arial', 12)), sg.FileBrowse(font=('Arial', 12))],
         [
          sg.Button('Gerar')]]
        self.janela = sg.Window('Métrica 90%').layout(layout)
        self.button, self.values = self.janela.Read()

    def Iniciar(self):
        self.janela.close()

class Loop:

    def Duplo(max_linha_raiz, max_coluna_referencia, max_linha_referencia, sheet_raiz, sheet_referencia, alinhamento, borda):
        for i in range(1, max_linha_raiz + 1):
            for j in range(1, max_coluna_referencia + 1):
                c = sheet_raiz.cell(row=i, column=j)
                if c.value == 'Data':
                    aux = i + 1
                    for n in range(aux, max_linha_raiz + 1):
                        for m in range(1, max_coluna_referencia + 1):
                            c = sheet_raiz.cell(row=n, column=m)
                            sheet_referencia.cell(row=(max_linha_referencia + 1), column=m).value = c.value
                            sheet_referencia.cell(row=(max_linha_referencia + 1), column=m).alignment = alinhamento
                            sheet_referencia.cell(row=(max_linha_referencia + 1), column=m).border = borda
                        max_linha_referencia += 1

    def Duplo2(linha_inicial, max_linha_raiz, coluna_inicial, coluna_final, max_linha_referencia, sheet_raiz, sheet_referencia, alinhamento, borda):
        for i in range(linha_inicial, max_linha_referencia + 1):
            for j in range(coluna_inicial, coluna_final + 1):
                c = sheet_raiz.cell(row=(max_linha_raiz + 1), column=j)
                c2 = sheet_referencia.cell(row=i, column=j)
                c.value = c2.value
                c.alignment = alinhamento
                c.border = borda
            max_linha_raiz += 1

    def Unico(coluna, linha_inicial, linha_maxima, planilha, alinhamento, borda, texto):
        for i in range(linha_inicial, linha_maxima + 1):
            cellref = planilha.cell(row=i, column=coluna)
            cellref.value = texto
            cellref.alignment = alinhamento
            cellref.border = borda

    def Condicional(coluna, linha_inicial, linha_maxima, planilha, alinhamento, borda, texto, texto2, color):
        for i in range(linha_inicial, linha_maxima + 1):
            celula = planilha.cell(row=i, column=coluna)
            if celula.value == texto2:
                celula.value = texto
                celula.alignment = alinhamento
                celula.border = borda
                celula.fill = color

    def Condicional2(coluna, coluna2, linha_inicial, linha_maxima, planilha, alinhamento, borda, texto, texto2, color):
        for i in range(linha_inicial, linha_maxima + 1):
            celula = planilha.cell(row=i, column=coluna)
            celula2 = planilha.cell(row=i, column=coluna2)
            if celula.value == texto:
                celula2.value = texto2
                celula2.alignment = alinhamento
                celula2.border = borda
                celula2.fill = color

    def Condicional3(coluna, coluna2, coluna3, linha_inicial, linha_maxima, planilha, alinhamento, borda, texto, texto2, texto3):
        for i in range(linha_inicial, linha_maxima + 1):
            celula = planilha.cell(row=i, column=coluna)
            celula2 = planilha.cell(row=i, column=coluna3)
            if celula.value == texto2 and celula2.value is not None and texto3 in celula2.value:
                cellref = planilha.cell(row=i, column=coluna2)
                cellref.value = texto
                cellref.alignment = alinhamento
                cellref.border = borda

    def Condicional4(coluna, coluna2, coluna3, linha_inicial, linha_maxima, planilha, alinhamento, borda, texto, texto2, texto3):
        for i in range(linha_inicial, linha_maxima + 1):
            celula = planilha.cell(row=i, column=coluna)
            celula2 = planilha.cell(row=i, column=coluna3)
            if celula.value == texto2 and celula2.value is not None and texto3 == celula2.value:
                cellref = planilha.cell(row=i, column=coluna2)
                cellref.value = texto
                cellref.alignment = alinhamento
                cellref.border = borda

    def CondicionalDuplo(coluna, coluna2, linha_inicial, linha_maxima, planilha, alinhamento, borda, texto, texto2, texto3, color):
        for i in range(linha_inicial, linha_maxima + 1):
            celula = planilha.cell(row=i, column=coluna)
            celula2 = planilha.cell(row=i, column=coluna2)
            if celula.value == texto2 and celula2.value == texto3:
                celula.value = texto
                celula.alignment = alinhamento
                celula.border = borda
                celula.fill = color

    def CondicionalDuploInverso(coluna, coluna2, linha_inicial, linha_maxima, planilha, alinhamento, borda, texto, texto2, texto3, color):
        for i in range(linha_inicial, linha_maxima + 1):
            celula = planilha.cell(row=i, column=coluna)
            celula2 = planilha.cell(row=i, column=coluna2)
            if celula.value == texto2 and celula2.value != texto3:
                celula2.value = texto
                celula2.alignment = alinhamento
                celula2.border = borda
                celula2.fill = color         
                
    def Deletar(coluna, linha_inicial, linha_maxima, planilha, texto2):
        i = linha_inicial
        while i <= linha_maxima:
            celula = planilha.cell(row=i, column=coluna)
            if celula.value is not None:
                if texto2 in celula.value:
                    planilha.delete_rows(i, 1)
                else:
                    i += 1
            else:
                i += 1

    def Deletar2(coluna, linha_inicial, linha_maxima, planilha, texto2):
        i = linha_inicial
        while i <= linha_maxima:
            celula = planilha.cell(row=i, column=coluna)
            if celula.value is not None:
                if celula.value == texto2:
                    planilha.delete_rows(i, 1)
                else:
                    i += 1
            else:
                i += 1

    def Deletar3(coluna, linha_inicial, linha_maxima, planilha, texto2):
        i = linha_inicial
        while i <= linha_maxima:
            celula = planilha.cell(row=i, column=coluna)
            if celula.value is not None:
                if celula.value != texto2:
                    planilha.delete_rows(i, 1)
                else:
                    i += 1
            else:
                i += 1

    def Deletar4(coluna, linha_inicial, linha_maxima, planilha, texto):
        i = linha_inicial
        while i <= linha_maxima:
            celula = planilha.cell(row=i, column=coluna)
            if celula.value is not None:
                if celula.value <= texto:
                    planilha.delete_rows(i, 1)
                else:
                    i += 1
            else:
                i += 1

    def VerificarMarcacao(linha_inicial, linha_maxima, sheet, equip, local, class_equip, local1, local2, color):
        wb5 = load_workbook('\\\\srvoffice\\PL\\PLM\\Coord_Movimentação_Equipamentos\\10. Power BI\\Logística Interna - Indicadores\\Acompanhamento Diário\\Relatório de Falhas de Operador.xlsx')
        ws5 = wb5['Sheet1']
        max_row_ws5 = ws5.max_row
        for i in range(linha_inicial, linha_maxima + 1):
            if sheet.cell(i, equip).value == class_equip and sheet.cell(i, local).value != local1 and sheet.cell(i, local).value != local2:
                for j in range(1, 9):
                    ws5.cell(max_row_ws5 + 1, j).value = sheet.cell(i, j).value
                ws5.cell(max_row_ws5 + 1, 9).value = 'Marcação de Área'
                max_row_ws5 = ws5.max_row
                sheet.cell(i, equip).fill = color
        wb5.save('\\\\srvoffice\\PL\\PLM\\Coord_Movimentação_Equipamentos\\10. Power BI\\Logística Interna - Indicadores\\Acompanhamento Diário\\Relatório de Falhas de Operador.xlsx')

    def VerificarMarcacao2(linha_inicial, linha_maxima, sheet, coluna1, coluna2, coluna3, class_equip, local, color, identificador):
        wb5 = load_workbook('\\\\srvoffice\\PL\\PLM\\Coord_Movimentação_Equipamentos\\10. Power BI\\Logística Interna - Indicadores\\Acompanhamento Diário\\Relatório de Falhas de Operador.xlsx')
        ws5 = wb5['Sheet1']
        max_row_ws5 = ws5.max_row
        for i in range(linha_inicial, linha_maxima + 1):
            if sheet.cell(i, coluna1).value != class_equip and sheet.cell(i, coluna2).value == local and sheet.cell(i, coluna3).value == identificador:
                for j in range(1, 9):
                    ws5.cell(max_row_ws5 + 1, j).value = sheet.cell(i, j).value
                ws5.cell(max_row_ws5 + 1, 9).value = 'Marcação de Área'
                max_row_ws5 = ws5.max_row
                sheet.cell(i, coluna1).fill = color
        wb5.save('\\\\srvoffice\\PL\\PLM\\Coord_Movimentação_Equipamentos\\10. Power BI\\Logística Interna - Indicadores\\Acompanhamento Diário\\Relatório de Falhas de Operador.xlsx')

    def VerificarMarcacao3(linha_inicial, linha_maxima, sheet, coluna1, coluna2, coluna3, turno, local, color, identificador):
        wb5 = load_workbook('\\\\srvoffice\\PL\\PLM\\Coord_Movimentação_Equipamentos\\10. Power BI\\Logística Interna - Indicadores\\Acompanhamento Diário\\Relatório de Falhas de Operador.xlsx')
        ws5 = wb5['Sheet1']
        max_row_ws5 = ws5.max_row
        for i in range(linha_inicial, linha_maxima + 1):
            if sheet.cell(i, coluna1).value != turno and sheet.cell(i, coluna2).value == local and sheet.cell(i, coluna3).value == identificador:
                for j in range(1, 9):
                    ws5.cell(max_row_ws5 + 1, j).value = sheet.cell(i, j).value
                ws5.cell(max_row_ws5 + 1, 9).value = 'Marcação de Turno'
                max_row_ws5 = ws5.max_row
                sheet.cell(i, coluna1).fill = color
        wb5.save('\\\\srvoffice\\PL\\PLM\\Coord_Movimentação_Equipamentos\\10. Power BI\\Logística Interna - Indicadores\\Acompanhamento Diário\\Relatório de Falhas de Operador.xlsx')

    def VerificarMarcacao4(linha_inicial, linha_maxima, sheet, coluna1, coluna2, coluna3, turno, local, color, identificador):
        wb5 = load_workbook('\\\\srvoffice\\PL\\PLM\\Coord_Movimentação_Equipamentos\\10. Power BI\\Logística Interna - Indicadores\\Acompanhamento Diário\\Relatório de Falhas de Operador.xlsx')
        ws5 = wb5['Sheet1']
        max_row_ws5 = ws5.max_row
        for i in range(linha_inicial, linha_maxima + 1):
            if sheet.cell(i, coluna1).value != turno and sheet.cell(i, coluna2).value == local and sheet.cell(i, coluna3).value == identificador:
                for j in range(1, 9):
                    ws5.cell(max_row_ws5 + 1, j).value = sheet.cell(i, j).value
                ws5.cell(max_row_ws5 + 1, 9).value = 'Marcação de Área'
                max_row_ws5 = ws5.max_row
                sheet.cell(i, coluna1).fill = color
        wb5.save('\\\\srvoffice\\PL\\PLM\\Coord_Movimentação_Equipamentos\\10. Power BI\\Logística Interna - Indicadores\\Acompanhamento Diário\\Relatório de Falhas de Operador.xlsx')