from time import sleep
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from Recursos import AcesseMeuSite
import os, pyperclip, pyautogui
from datetime import date, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pyperclip, pyautogui, os, datetime
from Recursos import Loop
from time import sleep

user = os.path.expanduser('~')

yesterday = date.today() - timedelta(days=1)
d1 = yesterday.strftime("%d/%m/%Y")

laranja = PatternFill('solid', fgColor='e3b973')

conta_jeceaba = 'VALLOUREC-JECEABA'
usuario = 'lucas.saraujo'
senha = 'Lucas@2022'

navegador = webdriver.Chrome()

navegador.get('https://web.pdo-vallourec.com.br/#/')
    
AcesseMeuSite.Logar(navegador, conta_jeceaba, usuario, senha)
sleep(15)
    
pyperclip.copy(d1)
sleep(1)
navegador.find_element_by_xpath('//*[@id="mat-input-3"]').send_keys(Keys.LEFT_CONTROL + 'a')
sleep(1)
pyautogui.hotkey('ctrl', 'v')
sleep(2)

pyperclip.copy(d1)
sleep(1)
navegador.find_element_by_xpath('//*[@id="mat-input-4"]').send_keys(Keys.LEFT_CONTROL + 'a')
sleep(1)
pyautogui.hotkey('ctrl', 'v')
sleep(2)

navegador.find_element_by_xpath('//*[@id="mat-expansion-panel-header-1"]').click()
    
navegador.find_element_by_xpath('//*[@id="cdk-accordion-child-1"]/div/form/button').click()
sleep(20)
    
navegador.close()

wb = load_workbook(filename=(user + '\\Downloads\\Relatorio_disponibilidade_contratual_resumido.xlsx'))

ws = wb['Sheet1']

max_row = ws.max_row
max_col = ws.max_column

dest_filename = user + '\\OneDrive - Vallourec\\Disponibilidade Emp16tons Diário.xlsx'

for i in range(1, max_row + 1):
    c = ws.cell(row=i, column=1)
    if c.value == 'EMPILHADEIRA 16 TON':
        for n in range(1, 8):
            ws.cell(i, n).fill = laranja

i = 21
while i <= max_row:
    celula = ws.cell(row=i, column=2)
    if celula.value is not None:
        if celula.value != 'EMPILHADEIRA 16 TON':
            ws.delete_rows(i, 1)
        else:
            i += 1
    else:
        i += 1

wb.save(filename=dest_filename)

os.remove(user + '\\Downloads\\Relatorio_disponibilidade_contratual_resumido.xlsx')