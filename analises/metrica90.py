from openpyxl import load_workbook
from openpyxl.styles import Alignment, Side, Border, Font, PatternFill
import datetime, os
from Recursos import Loop, TelaMetrica

user = os.path.expanduser('~')

center_align = Alignment(horizontal='center', vertical='center')
low_align = Alignment(horizontal='center', vertical='bottom')
box = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
box2 = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
transparente = PatternFill('none')
vermelho = PatternFill('solid', fgColor='e37373')
verde = PatternFill('solid', fgColor='6be07f')

format = '%H:%M:%S'

tela = TelaMetrica()

tela.Iniciar()

wb = load_workbook(tela.values['disponibilidade'])

wb2 = load_workbook(tela.values['ausencia'])

ws = wb['Sheet1']

ws2 = wb2['Sheet1']

ws_max_row = ws.max_row

ws2_max_row = ws2.max_row

ws_max_col = ws.max_column

ws2_max_col = ws2.max_column

for i in range (1, ws_max_row):
    if ws.cell(i, 1).value == "CADASTRO":
        linha_inicial = i+2
        break

ws.cell(linha_inicial-1, ws_max_col+1).value = "Mês"
ws.cell(linha_inicial-1, ws_max_col+1).alignment = low_align
ws.cell(linha_inicial-1, ws_max_col+1).border = box2
ws.cell(linha_inicial-1, ws_max_col+1).font = Font(bold=True)

for i in range(linha_inicial, ws_max_row + 1):
    ws.cell(i, ws_max_col+1).value = int(ws.cell(i, 1).value[3:-5])
    ws.cell(i, ws_max_col+1).alignment = center_align
    ws.cell(i, ws_max_col+1).border = box

ws_max_col = ws.max_column

ws.cell(linha_inicial-1, ws_max_col+1).value = "Manutenção TOTAL"
ws.cell(linha_inicial-1, ws_max_col+1).alignment = low_align
ws.cell(linha_inicial-1, ws_max_col+1).border = box2
ws.cell(linha_inicial-1, ws_max_col+1).font = Font(bold=True)

Loop.Condicional(5, 5, ws2_max_row, ws2, center_align, box, 'CENTRAL', 'NÃO IDENTIFICADO', transparente)

for i in range(linha_inicial, ws_max_row+1):
    data = ws.cell(i, 1).value
    classe = ws.cell(i, 2).value
    local = ws.cell(i, 3).value
    turno = ws.cell(i, 4).value
    celula = '00:00:00'
    timeList = [celula]
    mysum = datetime.timedelta()
    for j in range(5, ws2_max_row+1):
        data2 = ws2.cell(j, 1).value
        classe2 = ws2.cell(j, 6).value
        local2 = ws2.cell(j, 5).value
        turno2 = ws2.cell(j, 2).value
        tipo = ws2.cell(j, 8).value
        duracao = ws2.cell(j, ws2_max_col).value
        if data == data2 and classe == classe2 and local == local2 and turno == turno2 and tipo == "Manutenção":
            timeList.append(str(duracao))
    for n in timeList:
        (h, m, s) = n.split(':')
        d = datetime.timedelta(hours=int(h), minutes=int(m), seconds=int(s))
        mysum += d
    ws.cell(i, ws_max_col+1).value = mysum
    ws.cell(i, ws_max_col+1).alignment = center_align
    ws.cell(i, ws_max_col+1).border = box

ws_max_col = ws.max_column

ws.cell(linha_inicial-1, ws_max_col+1).value = "Ausência TOTAL"
ws.cell(linha_inicial-1, ws_max_col+1).alignment = low_align
ws.cell(linha_inicial-1, ws_max_col+1).border = box2
ws.cell(linha_inicial-1, ws_max_col+1).font = Font(bold=True)


for i in range(linha_inicial, ws_max_row + 1):
    timeList = [ws.cell(i, 19).value]
    timeList.append(str(ws.cell(i,ws_max_col).value))
    mysum = datetime.timedelta()
    for n in timeList:
        (h, m, s) = n.split(':')
        d = datetime.timedelta(hours=int(h), minutes=int(m), seconds=int(s))
        mysum += d
    ws.cell(i, ws_max_col+1).value = mysum
    ws.cell(i, ws_max_col+1).alignment = center_align
    ws.cell(i, ws_max_col+1).border = box

ws_max_col = ws.max_column

ws.cell(linha_inicial-1, ws_max_col+1).value = "Disponível Logado - Ausências"
ws.cell(linha_inicial-1, ws_max_col+1).alignment = low_align
ws.cell(linha_inicial-1, ws_max_col+1).border = box2
ws.cell(linha_inicial-1, ws_max_col+1).font = Font(bold=True)

for i in range(linha_inicial, ws_max_row + 1):
    ausencia = ws.cell(i,ws_max_col).value
    (h, m, s) = str(ws.cell(i,10).value).split(':')
    logado_total = datetime.timedelta(hours=int(h), minutes=int(m), seconds=int(s))
    if ausencia > logado_total:
        ws.cell(i, ws_max_col+1).value = datetime.timedelta(hours=0, minutes=0, seconds=0)
        ws.cell(i, ws_max_col+1).alignment = center_align
        ws.cell(i, ws_max_col+1).border = box
    else:
        ws.cell(i, ws_max_col+1).value = logado_total - ausencia
        ws.cell(i, ws_max_col+1).alignment = center_align
        ws.cell(i, ws_max_col+1).border = box

ws_max_col = ws.max_column 

ws.cell(linha_inicial-1, ws_max_col+1).value = "Contrato"
ws.cell(linha_inicial-1, ws_max_col+1).alignment = low_align
ws.cell(linha_inicial-1, ws_max_col+1).border = box2
ws.cell(linha_inicial-1, ws_max_col+1).font = Font(bold=True)

for i in range(linha_inicial, ws_max_row + 1):
    (h, m, s) = str(ws.cell(i,6).value).split(':')
    contrato = datetime.timedelta(hours=int(h), minutes=int(m), seconds=int(s))
    (h, m, s) = str(ws.cell(i,7).value).split(':')
    intervalo = datetime.timedelta(hours=int(h), minutes=int(m), seconds=int(s))
    ws.cell(i, ws_max_col+1).value = contrato - intervalo
    ws.cell(i, ws_max_col+1).alignment = center_align
    ws.cell(i, ws_max_col+1).border = box

ws_max_col = ws.max_column

ws.cell(linha_inicial-1, ws_max_col+1).value = "TX disponibilidade"
ws.cell(linha_inicial-1, ws_max_col+1).alignment = low_align
ws.cell(linha_inicial-1, ws_max_col+1).border = box2
ws.cell(linha_inicial-1, ws_max_col+1).font = Font(bold=True)

for i in range(linha_inicial, ws_max_row + 1):
    tx = ws.cell(i, ws_max_col-1).value / ws.cell(i, ws_max_col).value 
    if tx > 1:
        ws.cell(i, ws_max_col+1).value = 1
        ws.cell(i, ws_max_col+1).alignment = center_align
        ws.cell(i, ws_max_col+1).border = box
    else:
        ws.cell(i, ws_max_col+1).value = tx
        ws.cell(i, ws_max_col+1).alignment = center_align
        ws.cell(i, ws_max_col+1).border = box

ws_max_col = ws.max_column

ws.cell(linha_inicial-1, ws_max_col+1).value = "Crítico"
ws.cell(linha_inicial-1, ws_max_col+1).alignment = low_align
ws.cell(linha_inicial-1, ws_max_col+1).border = box2
ws.cell(linha_inicial-1, ws_max_col+1).font = Font(bold=True)

for i in range(linha_inicial, ws_max_row + 1):
    if ws.cell(i, 2).value == "RUCKER" and ws.cell(i, 3).value == "CENTRAL":
        ws.cell(i, ws_max_col + 1).value = "SIM"
        ws.cell(i, ws_max_col+1).alignment = center_align
        ws.cell(i, ws_max_col+1).border = box
    elif ws.cell(i, 2).value == "BROOK" and ws.cell(i, 3).value == "CENTRAL":
        ws.cell(i, ws_max_col + 1).value = "SIM"
        ws.cell(i, ws_max_col+1).alignment = center_align
        ws.cell(i, ws_max_col+1).border = box
    elif ws.cell(i, 2).value == "PÁ CARREGADEIRA" and ws.cell(i, 3).value == "PLATSUL":
        ws.cell(i, ws_max_col + 1).value = "SIM"
        ws.cell(i, ws_max_col+1).alignment = center_align
        ws.cell(i, ws_max_col+1).border = box
    elif ws.cell(i, 2).value == "PÁ CARREGADEIRA" and ws.cell(i, 3).value == "DPA":
        ws.cell(i, ws_max_col + 1).value = "SIM"
        ws.cell(i, ws_max_col+1).alignment = center_align
        ws.cell(i, ws_max_col+1).border = box
    elif ws.cell(i, 2).value == "PÁ CARREGADEIRA" and ws.cell(i, 3).value == "PATIOCRL":
        ws.cell(i, ws_max_col + 1).value = "SIM"
        ws.cell(i, ws_max_col+1).alignment = center_align
        ws.cell(i, ws_max_col+1).border = box
    elif ws.cell(i, 2).value == "EMPILHADEIRA" and ws.cell(i, 3).value == "PATIOCRL":
        ws.cell(i, ws_max_col + 1).value = "SIM"
        ws.cell(i, ws_max_col+1).alignment = center_align
        ws.cell(i, ws_max_col+1).border = box
    elif ws.cell(i, 2).value == "EMPILHADEIRA" and ws.cell(i, 3).value == "LUVAS":
        ws.cell(i, ws_max_col + 1).value = "SIM"
        ws.cell(i, ws_max_col+1).alignment = center_align
        ws.cell(i, ws_max_col+1).border = box
    elif ws.cell(i, 2).value == "EMPILHADEIRA" and ws.cell(i, 3).value == "FORJA MATÉRIA PRIMA":
        ws.cell(i, ws_max_col + 1).value = "SIM"
        ws.cell(i, ws_max_col+1).alignment = center_align
        ws.cell(i, ws_max_col+1).border = box
    elif ws.cell(i, 2).value == "PÁ CARREGADEIRA" and ws.cell(i, 3).value == "PÁTIO CALÇADO":
        ws.cell(i, ws_max_col + 1).value = "SIM"
        ws.cell(i, ws_max_col+1).alignment = center_align
        ws.cell(i, ws_max_col+1).border = box
    elif ws.cell(i, 2).value == "EMPILHADEIRA" and ws.cell(i, 3).value == "PÁTIO DE ASFALTO":
        ws.cell(i, ws_max_col + 1).value = "SIM"
        ws.cell(i, ws_max_col+1).alignment = center_align
        ws.cell(i, ws_max_col+1).border = box
    elif ws.cell(i, 2).value == "RUCKER" and ws.cell(i, 3).value == "PÁTIO CALÇADO":
        ws.cell(i, ws_max_col + 1).value = "SIM"
        ws.cell(i, ws_max_col+1).alignment = center_align
        ws.cell(i, ws_max_col+1).border = box
    else:
        ws.cell(i, ws_max_col + 1).value = "NÃO"
        ws.cell(i, ws_max_col+1).alignment = center_align
        ws.cell(i, ws_max_col+1).border = box

ws_max_col = ws.max_column

ws1 = wb.create_sheet("Sheet4")

ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

ws1['A1'].value = "TOTAL"
ws1['A1'].alignment = low_align
ws1['A2'].value = "100%"
ws1['A3'].value = "95%"
ws1['A4'].value = "90%"
ws1['A5'].value = "< 90%"
ws1['A2'].fill = verde
ws1['A3'].fill = verde
ws1['A4'].fill = verde
ws1['A5'].fill = vermelho

total = 0
cem_porcento = 0
noventa_e_cinco_porcento = 0
noventa_porcento = 0
menos_noventa_porcento = 0

for i in range(linha_inicial, ws_max_row+1):
    total += 1
    if ws.cell(i, ws_max_col-1).value == 1:
        cem_porcento += 1
    elif ws.cell(i, ws_max_col-1).value != 1 and ws.cell(i, ws_max_col-1).value >= 0.95:
        noventa_e_cinco_porcento += 1
    elif ws.cell(i, ws_max_col-1).value >= 0.90 and ws.cell(i, ws_max_col-1).value < 0.95:
        noventa_porcento += 1
    else:
        menos_noventa_porcento += 1

ws1['B2'].value = cem_porcento
ws1['B3'].value = noventa_e_cinco_porcento
ws1['B4'].value = noventa_porcento
ws1['B5'].value = menos_noventa_porcento
ws1['B6'].value = total
ws1['B2'].fill = verde
ws1['B3'].fill = verde
ws1['B4'].fill = verde
ws1['B5'].fill = vermelho

ws1['C2'].value = cem_porcento / total
ws1['C3'].value = noventa_e_cinco_porcento / total
ws1['C4'].value = noventa_porcento / total
ws1['C5'].value = menos_noventa_porcento / total
ws1['C6'].value = 1
ws1['C2'].fill = verde
ws1['C3'].fill = verde
ws1['C4'].fill = verde
ws1['C5'].fill = vermelho

ws1.merge_cells(start_row=8, start_column=1, end_row=8, end_column=3)

ws1['A8'].value = "NÃO CRÍTICO"
ws1['A8'].alignment = low_align
ws1['A9'].value = "100%"
ws1['A10'].value = "95%"
ws1['A11'].value = "90%"
ws1['A12'].value = "< 90%"
ws1['A9'].fill = verde
ws1['A10'].fill = verde
ws1['A11'].fill = verde
ws1['A12'].fill = vermelho

total = 0
cem_porcento = 0
noventa_e_cinco_porcento = 0
noventa_porcento = 0
menos_noventa_porcento = 0

for i in range(linha_inicial, ws_max_row+1):
    if ws.cell(i, ws_max_col).value == "NÃO":
        total += 1
    if ws.cell(i, ws_max_col-1).value == 1 and ws.cell(i, ws_max_col).value == "NÃO":
        cem_porcento += 1
    elif ws.cell(i, ws_max_col-1).value != 1 and ws.cell(i, ws_max_col-1).value >= 0.95 and ws.cell(i, ws_max_col).value == "NÃO":
        noventa_e_cinco_porcento += 1
    elif ws.cell(i, ws_max_col-1).value >= 0.90 and ws.cell(i, ws_max_col-1).value < 0.95 and ws.cell(i, ws_max_col).value == "NÃO":
        noventa_porcento += 1
    elif ws.cell(i, ws_max_col).value == "NÃO":
        menos_noventa_porcento += 1

ws1['B9'].value = cem_porcento
ws1['B10'].value = noventa_e_cinco_porcento
ws1['B11'].value = noventa_porcento
ws1['B12'].value = menos_noventa_porcento
ws1['B13'].value = total
ws1['B9'].fill = verde
ws1['B10'].fill = verde
ws1['B11'].fill = verde
ws1['B12'].fill = vermelho

ws1['C9'].value = cem_porcento / total
ws1['C10'].value = noventa_e_cinco_porcento / total
ws1['C11'].value = noventa_porcento / total
ws1['C12'].value = menos_noventa_porcento / total
ws1['C13'].value = 1
ws1['C9'].fill = verde
ws1['C10'].fill = verde
ws1['C11'].fill = verde
ws1['C12'].fill = vermelho

ws1.merge_cells(start_row=15, start_column=1, end_row=15, end_column=3)

ws1['A15'].value = "CRÍTICO"
ws1['A15'].alignment = low_align
ws1['A16'].value = "100%"
ws1['A17'].value = "95%"
ws1['A18'].value = "90%"
ws1['A19'].value = "< 90%"
ws1['A16'].fill = verde
ws1['A17'].fill = verde
ws1['A18'].fill = verde
ws1['A19'].fill = vermelho

total = 0
cem_porcento = 0
noventa_e_cinco_porcento = 0
noventa_porcento = 0
menos_noventa_porcento = 0

for i in range(linha_inicial, ws_max_row+1):
    if ws.cell(i, ws_max_col).value == "SIM":
        total += 1
    if ws.cell(i, ws_max_col-1).value == 1 and ws.cell(i, ws_max_col).value == "SIM":
        cem_porcento += 1
    elif ws.cell(i, ws_max_col-1).value != 1 and ws.cell(i, ws_max_col-1).value >= 0.95 and ws.cell(i, ws_max_col).value == "SIM":
        noventa_e_cinco_porcento += 1
    elif ws.cell(i, ws_max_col-1).value >= 0.90 and ws.cell(i, ws_max_col-1).value < 0.95 and ws.cell(i, ws_max_col).value == "SIM":
        noventa_porcento += 1
    elif ws.cell(i, ws_max_col).value == "SIM":
        menos_noventa_porcento += 1

ws1['B16'].value = cem_porcento
ws1['B17'].value = noventa_e_cinco_porcento
ws1['B18'].value = noventa_porcento
ws1['B19'].value = menos_noventa_porcento
ws1['B20'].value = total
ws1['B16'].fill = verde
ws1['B17'].fill = verde
ws1['B18'].fill = verde
ws1['B19'].fill = vermelho

ws1['C16'].value = cem_porcento / total
ws1['C17'].value = noventa_e_cinco_porcento / total
ws1['C18'].value = noventa_porcento / total
ws1['C19'].value = menos_noventa_porcento / total
ws1['C20'].value = 1
ws1['C16'].fill = verde
ws1['C17'].fill = verde
ws1['C18'].fill = verde
ws1['C19'].fill = vermelho

for i in range(linha_inicial, ws_max_row + 1):    
    ws.cell(i, 1).value = datetime.datetime.strptime(ws.cell(i, 1).value, '%d/%m/%Y').date()

wb.save(user + r'\Downloads\métrica90%.xlsx')
