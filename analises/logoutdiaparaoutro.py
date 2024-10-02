from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import datetime

user = os.path.expanduser('~')

dest_filename = user + '\\Downloads\\Resultado (Teste).xlsx'

wb2 = load_workbook(filename=(user + '\\Downloads\\teste.xlsx'))

vermelho = PatternFill('solid', fgColor='e37373')
roxo = PatternFill('solid', fgColor='9e58c7')

format = "%d/%m/%Y %H:%M:%S"

ws2 = wb2['Sheet1']

max_row2 = ws2.max_row
max_col2 = ws2.max_column

for i in range(5, max_row2 + 1):

    j = 5
    data_hora = datetime.datetime.strptime(ws2.cell(i, 10).value, format)

    while j <= max_row2:

        if ws2.cell(i, 8).value == ws2.cell(j, 8).value and ws2.cell(i, 5).value == ws2.cell(j, 5).value and ws2.cell(j, 1).value == ws2.cell(i, 1).value + datetime.timedelta(1):
            data_hora2 = datetime.datetime.strptime(ws2.cell(j, 9).value, format)
            if data_hora2.hour == data_hora.hour and data_hora2.minute >= data_hora.minute and data_hora2.minute < data_hora.minute + 30: 
                if ws2.cell(j, 13).value < datetime.time(1) and ws2.cell(j, 11).value > datetime.time(6):
                    ws2.cell(j, 11).fill = roxo
                    j +=1
                j +=1 
            j +=1 
        j += 1

wb2.save(filename=dest_filename)

#pyperclip.copy(user + '\\Downloads\\Resultado (Teste).xlsx')
#pyautogui.hotkey('win', 'r')
#sleep(1)
#pyautogui.hotkey('ctrl', 'v')
#sleep(1)
#pyautogui.press('enter')