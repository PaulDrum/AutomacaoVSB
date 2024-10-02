from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
from time import sleep
import pyperclip, pyautogui
import datetime

user = os.path.expanduser('~')

dest_filename = user + '\\Downloads\\Resultado (Teste).xlsx'

wb2 = load_workbook(filename=(user + '\\Downloads\\teste.xlsx'))

format = "%d/%m/%Y %H:%M:%S"

vermelho = PatternFill('solid', fgColor='e37373')

ws2 = wb2['Sheet1']

max_row2 = ws2.max_row
max_col2 = ws2.max_column

for m in range(5, max_row2 + 1):
    if ws2.cell(m, 11).value > datetime.time(12, 59, 59):
        ws2.cell(m, 11).value = datetime.time(12)
        ws2.cell(m, 8).fill = vermelho
    if ws2.cell(m, 12).value > datetime.time(12):
        ws2.cell(m, 12).value = datetime.time(12)
        ws2.cell(m, 8).fill = vermelho
    if ws2.cell(m, 13).value > datetime.time(12):
        ws2.cell(m, 13).value = datetime.time(12)
        ws2.cell(m, 8).fill = vermelho

for i in range(5, max_row2 + 1):
    
    j = 5
    timeList = [str(ws2.cell(i, 11).value)]
    mysum = datetime.timedelta()
    
    while j <= max_row2:
        
        if ws2.cell(i, 8).value == ws2.cell(j, 8).value and ws2.cell(i, 1).value == ws2.cell(j, 1).value and ws2.cell(i, 11).value != ws2.cell(j, 11).value and ws2.cell(i, 5).value == ws2.cell(j, 5).value:
            timeList.append(str(ws2.cell(j, 11).value))
            j += 1
        j+=1
    
    for n in timeList:
        (h, m, s) = n.split(':')
        d = datetime.timedelta(hours=int(h), minutes=int(m), seconds=int(s))
        mysum += d

    if mysum > datetime.timedelta(hours=12, minutes=59, seconds=59):
        ws2.cell(i, 8).fill = vermelho


for i in range(5, max_row2 + 1):

    j = 5
    data_hora = datetime.datetime.strptime(ws2.cell(i, 10).value, format)
    timeList = [str(ws2.cell(i, 11).value)]
    mysum = datetime.timedelta()
    ws2.cell(i, 1).value = datetime.datetime.strptime(ws2.cell(i, 1).value, '%d/%m/%Y').date()

    while j <= max_row2:
        if ws2.cell(i, 8).value == ws2.cell(j, 8).value and ws2.cell(i, 5).value != ws2.cell(j, 5).value and ws2.cell(j, 1).value == ws2.cell(i, 1).value + datetime.timedelta(1) or ws2.cell(j, 1).value == ws2.cell(i, 1).value and ws2.cell(i, 8).value == ws2.cell(j, 8).value and ws2.cell(i, 5).value != ws2.cell(j, 5).value:
            data_hora2 = datetime.datetime.strptime(ws2.cell(j, 9).value, format)
            if data_hora2.hour == data_hora.hour and data_hora2.minute >= data_hora.minute and data_hora2.minute < data_hora.minute + 30: 
                timeList.append(str(ws2.cell(j, 11).value))
                j += 1
            j+=1
        j += 1
    
    for n in timeList:
        (h, m, s) = n.split(':')
        d = datetime.timedelta(hours=int(h), minutes=int(m), seconds=int(s))
        mysum += d

    if mysum > datetime.timedelta(hours=12, minutes=59, seconds=59):
        ws2.cell(i, 8).fill = vermelho

wb2.save(filename=dest_filename)

pyperclip.copy(user + '\\Downloads\\Resultado (Teste).xlsx')
pyautogui.hotkey('win', 'r')
sleep(1)
pyautogui.hotkey('ctrl', 'v')
sleep(1)
pyautogui.press('enter')