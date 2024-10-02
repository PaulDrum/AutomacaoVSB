from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
from time import sleep
import pyperclip, pyautogui

user = os.path.expanduser('~')

dest_filename = user + '\\Downloads\\Resultado (Teste).xlsx'

wb2 = load_workbook(filename=(user + '\\Downloads\\teste.xlsx'))

vermelho = PatternFill('solid', fgColor='e37373')

ws2 = wb2['Sheet1']

max_row2 = ws2.max_row
max_col2 = ws2.max_column

for i in range(5, max_row2 + 1):
    j = 5
    while j <= max_row2:
        if ws2.cell(i, 8).value == ws2.cell(j, 8).value and ws2.cell(i, 1).value == ws2.cell(j, 1).value and ws2.cell(i, 5).value != ws2.cell(j, 5).value and ws2.cell(i, 3).value == ws2.cell(j, 3).value:
            if ws2.cell(i, 9).value < ws2.cell(j, 9).value and ws2.cell(i, 10).value > ws2.cell(j, 9).value:
                if ws2.cell(i, 13).value > ws2.cell(j, 13).value:
                    ws2.delete_rows(j, 1)
                    ws2.cell(i, 8).fill = vermelho
                else:
                    ws2.delete_rows(i, 1)
                    ws2.cell(j, 8).fill = vermelho
            else:
                j += 1
        else:
            j += 1

wb2.save(filename=dest_filename)

pyperclip.copy(user + '\\Downloads\\Resultado (Teste).xlsx')
pyautogui.hotkey('win', 'r')
sleep(1)
pyautogui.hotkey('ctrl', 'v')
sleep(1)
pyautogui.press('enter')