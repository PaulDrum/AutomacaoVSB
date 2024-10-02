from openpyxl import load_workbook

wb = load_workbook(r'C:\\Users\\s-Lucas.SAraujo\\Downloads\\planilha_combinada.xlsx')

ws = wb.active

max_row = ws.max_row

ws.cell(1, 9).value = "Nº"

for i in range(1, max_row+1):
    array = []
    idx=1
    idj=1
    if ws.cell(i, 4).value == "12 x 36 Noite" or ws.cell(i, 4).value == "N2" and ws.cell(i, 9).value is None:
        array.append(ws.cell(i, 1).value)
        array.append(ws.cell(i, 2).value)
        array.append(ws.cell(i, 3).value)
        array.append(ws.cell(i, 4).value)
        array.append(ws.cell(i, 8).value)
        ws.cell(i, 9).value = idx
        for j in range(1, max_row+1):
            array2 = []
            if i == j:
                pass
            elif ws.cell(j, 4).value == "12 x 36 Noite" or ws.cell(j, 4).value == "N2":        
                array2.append(ws.cell(j, 1).value)
                array2.append(ws.cell(j, 2).value)
                array2.append(ws.cell(j, 3).value)
                array2.append(ws.cell(j, 4).value)
                array2.append(ws.cell(j, 8).value)
                if array == array2 and ws.cell(i, 5).value == ws.cell(j, 5).value:
                    idx+=1
                    ws.cell(j, 9).value = idx  
                elif array == array2 and ws.cell(i, 5).value != ws.cell(j, 5).value:
                    ws.cell(j, 9).value = idj
                    idj+=1
                else:
                    pass
                    
    else:
        pass

wb.save(r'C:\\Users\\s-Lucas.SAraujo\\Downloads\\TesteLogicaDoida.xlsx')
    
            