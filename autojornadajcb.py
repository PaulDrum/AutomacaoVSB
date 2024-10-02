from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from suporte.Recursos import AcesseMeuSite, TelaJornadas
import pyautogui, pyperclip
from openpyxl import load_workbook

tela = TelaJornadas()

tela

if tela.values['6x1'] or tela.values['12x36'] == True:

    conta = "VALLOUREC-JECEABA"

    Turnos = ['Turno 1', 'Turno 2', 'Turno 3', 'Turno Novo 1', 'Turno Novo 2']

    intervalo = ['60', '0']

    # chrome_options = webdriver.EdgeOptions()
    # chrome_options.add_argument('--start-maximized')
    # chrome_options.add_argument("--disable-infobars")
    # chrome_options.add_experimental_option("useAutomationExtension", False)
    # chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    # chrome_options.add_argument("--window-position=0,0")
    # chrome_options.add_experimental_option("detach", True)

    navegador = webdriver.Edge(executable_path=r'C:\Users\s-Lucas.SAraujo\AppData\Local\Programs\Python\Python36-32\msedgedriver.exe')

    sleep(5)
    
    navegador.switch_to.window(navegador.window_handles[0])

    navegador.get("https://web.pdo-vallourec.com.br/#/")

    sleep(5)

    AcesseMeuSite.Logar(navegador, conta, tela.values['usuario'], tela.values['senha'])
    
    sleep(22)

    navegador.find_element(By.XPATH, "//*[contains(text(),'Programação de Jornadas')]").click()

    sleep(5)

    wb = load_workbook(tela.values['doc'])

if tela.values['6x1'] == True:

    ws = wb['Jeceaba']

    max_row_ws = ws.max_row

    MatrizTurnos = ["T1", "T2", "T3", "T1", "T2", "T3"]

    for i in range(2, max_row_ws+1):
        for j in range(5, 11):
            if ws.cell(i, j).value is not None:
                MatrizTurnos[j - 5] = ws.cell(i, j).value
            else:
                MatrizTurnos[j - 5] = 0

        for n in range(0, 6):

            if n == 3 and MatrizTurnos[0] != 0:
                continue

            if n == 4 and MatrizTurnos[1] != 0:
                continue

            if n == 5 and MatrizTurnos[2] != 0:
                continue

            if MatrizTurnos[n] != 0:
                    
                qtdJornadas = MatrizTurnos[n]

                navegador.find_element(By.XPATH, "//*[contains(text(),'Nova programação de jornada')]").click()
                sleep(1)   

                # testando datas
                pyperclip.copy(tela.values['-CAL-'])
                sleep(0.5)
                navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[1]/expert-date-picker[1]/mat-form-field/div/div[1]/div[1]/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                sleep(0.5)
                pyautogui.hotkey('ctrl', 'v')
                sleep(0.5)
                pyperclip.copy(tela.values['-CAL2-'])
                sleep(0.5)
                navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[1]/expert-date-picker[2]/mat-form-field/div/div[1]/div[1]/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                sleep(0.5)
                pyautogui.hotkey('ctrl', 'v')
                sleep(0.5)

                # testando identificador
                navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[2]/mat-form-field[1]/div/div[1]/div/mat-select/div/div[1]').click()
                sleep(0.5)
                navegador.find_element(By.XPATH, "//*[contains(text(),'Jeceaba-Tradimaq')]").click()
                sleep(0.5)

                if ws.cell(i, 3).value == "SEG/SAB":
                    navegador.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[4]/mat-form-field[2]/div/div[1]/div/mat-select/div/div[1]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, "//*[contains(text(),'Segunda-feira')]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, "//*[contains(text(),'Terça-feira')]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, "//*[contains(text(),'Quarta-feira')]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, "//*[contains(text(),'Quinta-feira')]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, "//*[contains(text(),'Sexta-feira')]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, "//*[contains(text(),'Sábado')]").click()
                    sleep(0.5)
                    pyautogui.doubleClick(x=1175, y=762)
                    sleep(0.5)
                
                elif ws.cell(i, 3).value == "SEG/SEX":
                    navegador.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[4]/mat-form-field[2]/div/div[1]/div/mat-select/div/div[1]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, "//*[contains(text(),'Segunda-feira')]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, "//*[contains(text(),'Terça-feira')]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, "//*[contains(text(),'Quarta-feira')]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, "//*[contains(text(),'Quinta-feira')]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, "//*[contains(text(),'Sexta-feira')]").click()
                    sleep(0.5)
                    pyautogui.doubleClick(x=1175, y=762)
                    sleep(0.5)

                elif ws.cell(i, 3).value == "DOM/SEXTA":
                    navegador.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[4]/mat-form-field[2]/div/div[1]/div/mat-select/div/div[1]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, "//*[contains(text(),'Domingo')]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, "//*[contains(text(),'Segunda-feira')]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, "//*[contains(text(),'Terça-feira')]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, "//*[contains(text(),'Quarta-feira')]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, "//*[contains(text(),'Quinta-feira')]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, "//*[contains(text(),'Sexta-feira')]").click()
                    sleep(0.5)
                    pyautogui.doubleClick(x=1175, y=762)
                    sleep(0.5)

                else:
                    navegador.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[4]/mat-form-field[2]/div/div[1]/div/mat-select/div/div[1]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, "//*[contains(text(),'--Todos--')]").click()
                    sleep(0.5)
                    pyautogui.doubleClick(x=1175, y=762)
                    sleep(0.5)

                classe = ws.cell(i, 1).value
                
                navegador.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[2]/mat-form-field[2]/div/div[1]/div/mat-select/div/div[1]").click()
                sleep(0.5)
                navegador.find_element(By.XPATH, f"//*[contains(text(),'{classe}')]").click()
                sleep(0.5)
                
                area = ws.cell(i, 2).value

                navegador.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[5]/mat-form-field/div/div[1]/div/mat-select/div/div[1]").click()
                sleep(0.5)
                navegador.find_element(By.XPATH, f"//*[contains(text(),'{area}')]").click()
                sleep(0.5)

                if ws.cell(i, 4).value == "SIM":

                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[4]/div/div[1]/div/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[4]/div/div[1]/div/input').send_keys(intervalo)
                    sleep(0.5)

                #testando qtd_jornadas
                navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[4]/mat-form-field[1]/div/div[1]/div/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                sleep(0.5)
                navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[4]/mat-form-field[1]/div/div[1]/div/input').send_keys(qtdJornadas)
                sleep(0.5)

                if n == 0 or n == 3:
                        
                    #testando turno
                    navegador.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[1]/div/div[1]/div/mat-select/div/div[1]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, f"//*[contains(text(), '{Turnos[0]}')]").click()
                    sleep(0.5)

                    #testando horas
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[3]/div/div[1]/div/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[3]/div/div[1]/div/input').send_keys(tela.values["t1fim"])
                    sleep(0.5)

                    navegador.find_element(By.XPATH, '//*[contains(text(), "Salvar")]').click()

                elif n == 1 or n == 4:

                    #testando turno
                    navegador.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[1]/div/div[1]/div/mat-select/div/div[1]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, f"//*[contains(text(), '{Turnos[1]}')]").click()
                    sleep(0.5)

                    #testando horas
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[2]/div/div[1]/div/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[2]/div/div[1]/div/input').send_keys(tela.values["t2inicio"])
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[3]/div/div[1]/div/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[3]/div/div[1]/div/input').send_keys(tela.values["t2fim"])
                    sleep(0.5)
                    
                    navegador.find_element(By.XPATH, '//*[contains(text(), "Salvar")]').click()

                else:

                    #testando turno
                    navegador.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[1]/div/div[1]/div/mat-select/div/div[1]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, f"//*[contains(text(), '{Turnos[2]}')]").click()
                    sleep(0.5)

                    #testando horas
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[2]/div/div[1]/div/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[2]/div/div[1]/div/input').send_keys(tela.values["t3inicio"])
                    sleep(0.5)

                    navegador.find_element(By.XPATH, '//*[contains(text(), "Salvar")]').click()

                sleep(2)      

if tela.values['12x36'] == True:   

    ws = wb['Jeceaba 12x36']

    max_row_ws = ws.max_row

    MatrizTurnos = ['S1', 'S2', 'D1', 'D2']

    for i in range(2, max_row_ws+1):

        for j in range(5, 9):
            if ws.cell(i, j).value is not None:
                MatrizTurnos[j - 5] = ws.cell(i, j).value
            else:
                MatrizTurnos[j - 5] = 0

        for n in range(0, 4):

            if n == 2 and  MatrizTurnos[0] != 0:
                continue

            if n == 3 and  MatrizTurnos[1] != 0:
                continue

            if MatrizTurnos[n] != 0:

                if n == 1 or n == 3:

                    qtdJornadas = MatrizTurnos[n]

                    navegador.find_element(By.XPATH, "//*[contains(text(),'Nova programação de jornada')]").click()
                    sleep(1)   

                    # testando datas
                    pyperclip.copy(tela.values['-CAL-'])
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[1]/expert-date-picker[1]/mat-form-field/div/div[1]/div[1]/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.5)
                    pyautogui.hotkey('ctrl', 'v')
                    sleep(0.5)
                    pyperclip.copy(tela.values['-CAL2-'])
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[1]/expert-date-picker[2]/mat-form-field/div/div[1]/div[1]/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.5)
                    pyautogui.hotkey('ctrl', 'v')
                    sleep(0.5)

                    # testando identificador
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[2]/mat-form-field[1]/div/div[1]/div/mat-select/div/div[1]').click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, "//*[contains(text(),'Jeceaba-Tradimaq')]").click()
                    sleep(0.5)

                    navegador.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[4]/mat-form-field[2]/div/div[1]/div/mat-select/div/div[1]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, "//*[contains(text(),'--Todos--')]").click()
                    sleep(0.5)
                    pyautogui.doubleClick(x=1175, y=762)
                    sleep(0.5)

                    classe = ws.cell(i, 1).value
                    
                    navegador.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[2]/mat-form-field[2]/div/div[1]/div/mat-select/div/div[1]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, f"//*[contains(text(),'{classe}')]").click()
                    sleep(0.5)
                    
                    area = ws.cell(i, 2).value

                    navegador.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[5]/mat-form-field/div/div[1]/div/mat-select/div/div[1]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, f"//*[text()='{area}']").click()
                    sleep(0.5)

                    if ws.cell(i, 4).value == "SIM":
                        navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[4]/div/div[1]/div/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                        sleep(0.5)
                        navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[4]/div/div[1]/div/input').send_keys(intervalo[0])
                        sleep(0.5)
                    else:
                        navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[4]/div/div[1]/div/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                        sleep(0.5)
                        navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[4]/div/div[1]/div/input').send_keys(intervalo[1])
                        sleep(0.5)

                    #testando qtd_jornadas
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[4]/mat-form-field[1]/div/div[1]/div/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[4]/mat-form-field[1]/div/div[1]/div/input').send_keys(qtdJornadas)
                    sleep(0.5)
    
                    #testando turno
                    navegador.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[1]/div/div[1]/div/mat-select/div/div[1]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, f"//*[contains(text(), '{Turnos[3]}')]").click()
                    sleep(0.5)

                    #testando horas
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[2]/div/div[1]/div/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[2]/div/div[1]/div/input').send_keys(tela.values["12x36dinicio"])
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[3]/div/div[1]/div/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[3]/div/div[1]/div/input').send_keys(tela.values["12x36dfim"])
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '//*[contains(text(), "Salvar")]').click()

                    sleep(2) 

                else:    
                                
                    qtdJornadas = MatrizTurnos[n]

                    navegador.find_element(By.XPATH, "//*[contains(text(),'Nova programação de jornada')]").click()
                    sleep(1)   

                    # testando datas
                    pyperclip.copy(tela.values['-CAL-'])
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[1]/expert-date-picker[1]/mat-form-field/div/div[1]/div[1]/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.5)
                    pyautogui.hotkey('ctrl', 'v')
                    sleep(0.5)
                    pyperclip.copy(tela.values['-CAL2-'])
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[1]/expert-date-picker[2]/mat-form-field/div/div[1]/div[1]/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.5)
                    pyautogui.hotkey('ctrl', 'v')
                    sleep(0.5)
                    
                    # testando identificador
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[2]/mat-form-field[1]/div/div[1]/div/mat-select/div/div[1]').click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, "//*[contains(text(),'Jeceaba-Tradimaq')]").click()
                    sleep(0.5)

                    navegador.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[4]/mat-form-field[2]/div/div[1]/div/mat-select/div/div[1]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, "//*[contains(text(),'--Todos--')]").click()
                    sleep(0.5)
                    pyautogui.doubleClick(x=1175, y=762)
                    sleep(0.5)

                    classe = ws.cell(i, 1).value
                    
                    navegador.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[2]/mat-form-field[2]/div/div[1]/div/mat-select/div/div[1]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, f"//*[contains(text(),'{classe}')]").click()
                    sleep(0.5)
                    
                    area = ws.cell(i, 2).value

                    navegador.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[5]/mat-form-field/div/div[1]/div/mat-select/div/div[1]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, f"//*[text()='{area}']").click()
                    sleep(0.5)

                    #testando qtd_jornadas
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[4]/mat-form-field[1]/div/div[1]/div/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[4]/mat-form-field[1]/div/div[1]/div/input').send_keys(qtdJornadas)
                    sleep(0.5)
    
                    #testando turno
                    navegador.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[1]/div/div[1]/div/mat-select/div/div[1]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, f"//*[contains(text(), '{Turnos[4]}')]").click()
                    sleep(0.5)

                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[4]/div/div[1]/div/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[4]/div/div[1]/div/input').send_keys('0')
                    sleep(0.5)

                    #testando horas
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[2]/div/div[1]/div/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[2]/div/div[1]/div/input').send_keys(tela.values["12x36ninicio"])
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[3]/div/div[1]/div/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[3]/div/div[1]/div/input').send_keys('23:59:59')
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '//*[contains(text(), "Salvar")]').click()

                    sleep(2)   

                    navegador.find_element(By.XPATH, "//*[contains(text(),'Nova programação de jornada')]").click()
                    sleep(1)   

                    # testando datas
                    pyperclip.copy(tela.values['-CAL-'])
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[1]/expert-date-picker[1]/mat-form-field/div/div[1]/div[1]/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.5)
                    pyautogui.hotkey('ctrl', 'v')
                    sleep(0.5)
                    pyperclip.copy(tela.values['-CAL2-'])
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[1]/expert-date-picker[2]/mat-form-field/div/div[1]/div[1]/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.5)
                    pyautogui.hotkey('ctrl', 'v')
                    sleep(0.5)

                    # testando identificador
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[2]/mat-form-field[1]/div/div[1]/div/mat-select/div/div[1]').click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, "//*[contains(text(),'Jeceaba-Tradimaq')]").click()
                    sleep(0.5)

                    navegador.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[4]/mat-form-field[2]/div/div[1]/div/mat-select/div/div[1]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, "//*[contains(text(),'--Todos--')]").click()
                    sleep(0.5)
                    pyautogui.doubleClick(x=1175, y=762)
                    sleep(0.5)

                    classe = ws.cell(i, 1).value
                    
                    navegador.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[2]/mat-form-field[2]/div/div[1]/div/mat-select/div/div[1]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, f"//*[contains(text(),'{classe}')]").click()
                    sleep(0.5)
                    
                    area = ws.cell(i, 2).value

                    navegador.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[5]/mat-form-field/div/div[1]/div/mat-select/div/div[1]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, f"//*[text()='{area}']").click()
                    sleep(0.5)

                    if ws.cell(i, 4).value == "SIM":
                        navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[4]/div/div[1]/div/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                        sleep(0.5)
                        navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[4]/div/div[1]/div/input').send_keys(intervalo[0])
                        sleep(0.5)
                    else:
                        navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[4]/div/div[1]/div/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                        sleep(0.5)
                        navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[4]/div/div[1]/div/input').send_keys(intervalo[1])
                        sleep(0.5)

                    #testando qtd_jornadas
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[4]/mat-form-field[1]/div/div[1]/div/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[4]/mat-form-field[1]/div/div[1]/div/input').send_keys(qtdJornadas)
                    sleep(0.5)
    
                    #testando turno
                    navegador.find_element(By.XPATH, "/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[1]/div/div[1]/div/mat-select/div/div[1]").click()
                    sleep(0.5)
                    navegador.find_element(By.XPATH, f"//*[contains(text(), '{Turnos[4]}')]").click()
                    sleep(0.5)

                    #testando horas
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[2]/div/div[1]/div/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[2]/div/div[1]/div/input').send_keys('00:00:00')
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[3]/div/div[1]/div/input').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/mat-dialog-container/dialog-cadastra-programacao-jornada/mat-dialog-content/form-programacao-jornada/form/div[3]/mat-form-field[3]/div/div[1]/div/input').send_keys(tela.values["12x36nfim"])
                    sleep(0.5)
                    navegador.find_element(By.XPATH, '//*[contains(text(), "Salvar")]').click()

                    sleep(2)              