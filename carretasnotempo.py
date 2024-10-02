from openpyxl import load_workbook
from openpyxl.styles import Alignment, Side, Border, Font, PatternFill
import pyperclip, pyautogui, os, datetime
from suporte.Recursos import TelaMesclar
from suporte.Recursos import Loop
from time import sleep
import pandas as pd
from datetime import datetime
import xlsxwriter

user = os.path.expanduser('~')

data_buscada = "23/09/2024"

faixas_horas = ["00:00:00", ]

verde = PatternFill('solid', fgColor='6be07f')
amarelo = PatternFill('solid', fgColor='efee9b')

basewb = load_workbook(filename=(user + '\\Downloads\\Base Python.xlsx'),data_only=True)
destinowb = load_workbook(filename=(user + '\\Downloads\\Base Visual Carreta Hora a Hora.xlsx'))
    
dest_filename = user + '\\Downloads\\Resultado (Teste).xlsx'
    
basews = basewb['Sheet1']
destinows = destinowb['Tabela Visualização']
processosws = destinowb['Processos']
            
base_max_row = basews.max_row
base_max_col = basews.max_column

destino_max_row = destinows.max_row
destino_max_col = destinows.max_column

processos_max_row = processosws.max_row
processos_max_row = processosws.max_column

# Converter célula em DATETIME
# datetime.strptime(basews.cell(i, 7).value, "%d/%m/%Y %H:%M:%S")

for i in range(2, base_max_row+1):
    array_destino_atual = []
    area_atual = basews.cell(i, 5).value
    carreta_atual = basews.cell(i, 4).value
    chegadanodestino_atual = basews.cell(i, 2).value
    fimdoatendimento_atual = basews.cell(i, 3).value
    saidadaarea_atual = basews.cell(i, 7).value
    for n in range(2, processos_max_row+1):
        print(processosws.cell(n, 2).value)
        if area_atual == processosws.cell(n, 2).value:
            print(area_atual)
            array_destino_atual.append(processosws.cell(n, 1).value)
            array_destino_atual.append(processosws.cell(n, 2).value)
            break
    print(array_destino_atual)
    for y in range(1, destino_max_row+1):
        if destinows.cell(y,1).value is None:
            destinows.cell(y,1).value = array_destino_atual[0]
            destinows.cell(y,2).value = array_destino_atual[1]
    
destinowb.save(filename=dest_filename)