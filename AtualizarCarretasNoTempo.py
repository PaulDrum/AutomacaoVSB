from time import sleep
from selenium import webdriver
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Side, Border
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from suporte.Recursos import AcesseMeuSiteAnalitico
import os, pyperclip, pyautogui, datetime
from suporte.Recursos import TelaCarretas
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.user_credential import UserCredential

tela = TelaCarretas()

tela.Iniciar()

if tela.button == "Atualizar":

    user = os.path.expanduser('~')

    # Define o caminho do arquivo
    file_path = os.path.join(os.path.expanduser('~'), 'Downloads', 'Relatorio_viagens_carretinhas_ruckers.xlsx')

    # Verifica se o arquivo existe
    if os.path.exists(file_path):
        # Remove o arquivo
        os.remove(file_path)
    else:
        pass

    ## BUSCAR NO NAVEGADOR

    service = Service(EdgeChromiumDriverManager().install())
    navegador = webdriver.Edge(service=service)

    navegador.get("https://analitico-vallourec.com.br/login/")

    sleep(3)
        
    AcesseMeuSiteAnalitico.Logar(navegador, "lucas.saraujo", "Lucas@2022")

    sleep(2)

    navegador.find_element(By.XPATH, '//a[@data-toggle="sidebar"]').click()

    sleep(1)

    navegador.find_element(By.XPATH, '//a[@href="/relatorios/"]').click()

    sleep(2)

    pyperclip.copy(tela.values['-CAL-'])

    sleep(1)

    navegador.find_element(By.XPATH, '//input[@ng-model="data_inicial"]').send_keys(Keys.LEFT_CONTROL + "a")

    sleep(1)

    pyautogui.hotkey("ctrl","v")

    sleep(2)

    pyperclip.copy(tela.values['-CAL2-'])

    sleep(1)

    navegador.find_element(By.XPATH, '//input[@ng-model="data_final"]').send_keys(Keys.LEFT_CONTROL + "a")

    sleep(1)

    pyautogui.hotkey("ctrl","v")

    sleep(1)

    navegador.execute_script("window.scrollTo(0, 420)")

    sleep(1)

    navegador.find_element(By.XPATH, '//div[@heading="Relatório de Viagens de Carretinhas e Ruckers"]').click()

    sleep(1)

    navegador.find_element(By.XPATH, '//button[@ng-click="criarRelatorio(13);"]').click()

    sleep(42)

    navegador.quit()

    # Defina o caminho dos arquivos
    user = os.path.expanduser("~")  # Certifique-se de definir o caminho do usuário
    origem_path = os.path.join(user, 'Downloads', 'Relatorio_viagens_carretinhas_ruckers.xlsx')
    destino_path = r'\\Srvoffice\pl\PLM\Coord_Movimentação_Equipamentos\10. Power BI\Logística Interna - Indicadores\Acompanhamento Diário\Base Carretas.xlsx'

    # Carregue as planilhas
    basewb = load_workbook(filename=origem_path)
    destinowb = load_workbook(filename=destino_path)

    origemws = basewb['Sheet1']
    destinows = destinowb['Sheet1']

    # Defina alinhamento e borda
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Função para remover colunas desnecessárias
    def remove_unnecessary_columns(ws):
        # Deletar colunas de A a N (1 a 14)
        ws.delete_cols(1, 14)
        
        # Deletar coluna B (agora é a coluna A após a remoção anterior)
        ws.delete_cols(2, 1)  # Deleta a nova coluna B

        # Deletar colunas C e D (agora são as colunas B e C após as remoções anteriores)
        ws.delete_cols(3, 2)  # Deleta as novas colunas B e C

        # Deletar colunas D ao K (agora são as colunas B ao J após as remoções anteriores)
        ws.delete_cols(4, 8)  # Deleta as novas colunas B até J

        # Deletar colunas E ao L (agora são as colunas B ao K após as remoções anteriores)
        ws.delete_cols(5, 8)  # Deleta as novas colunas B até H
        
        # Deletar tudo após a coluna E (agora é a coluna A após todas as remoções anteriores)
        max_col = ws.max_column
        if max_col > 5:  # Se houver mais de 5 colunas
            ws.delete_cols(6, max_col - 5)  # Deleta da coluna F até a última

    # Remover colunas desnecessárias na planilha de origem
    remove_unnecessary_columns(origemws)

    # Converta datas na planilha de origem
    for i in range(1, origemws.max_row + 1):
        if origemws.cell(i, 1).value == 'Data':
            aux = i + 1
            for n in range(aux, origemws.max_row + 1):
                origemws.cell(n, 1).value = datetime.datetime.strptime(origemws.cell(n, 1).value, '%d/%m/%Y').date()
            break

    # Remove linhas em que a área destino é ---
    for i in range(origemws.max_row, 4, -1):  # Itera de trás para frente
        if origemws.cell(i, 4).value == "---":
            origemws.delete_rows(i)
    
    # Remove linhas em que a carreta é teste ddmx ---
    for i in range(origemws.max_row, 4, -1):  # Itera de trás para frente
        if origemws.cell(i, 3).value == "TESTE DDMX - NÃO USAR":
            origemws.delete_rows(i)

    # Verifica se há dados na planilha de origem após remoção
    print(f"Linha máxima após remoção: {origemws.max_row}")

    # Encontre a última linha da tabela existente na planilha de destino
    linha_maxima = destinows.max_row + 1  # A primeira linha livre após a última linha existente

    coluna_maxima = origemws.max_column

    # Copie os dados da planilha de origem para a planilha de destino
    for i in range(5, origemws.max_row + 1):
        for j in range(1, coluna_maxima + 1):
            destinows.cell(row=linha_maxima, column=j).value = origemws.cell(row=i, column=j).value
            destinows.cell(row=linha_maxima, column=j).alignment = center_align
            destinows.cell(row=linha_maxima, column=j).border = thin_border
        linha_maxima += 1

    # Remover linhas duplicadas na planilha de destino
    def remove_duplicates(ws):
        seen = set()
        rows_to_delete = []
        
        for row in ws.iter_rows(min_row=2):  # Começa da segunda linha para ignorar cabeçalho
            row_data = tuple(cell.value for cell in row)  # Cria uma tupla com os valores da linha
            
            if row_data in seen:
                rows_to_delete.append(row[0].row)  # Adiciona a linha à lista de exclusão
            else:
                seen.add(row_data)  # Adiciona a tupla ao conjunto
        
        for row_index in sorted(rows_to_delete, reverse=True):  # Remove linhas do fundo para o topo
            ws.delete_rows(row_index)

    remove_duplicates(destinows)  # Chama a função para remover duplicatas

    # Salve a planilha de destino atualizada
    destinowb.save(filename=destino_path)
    print("Planilha salva com sucesso.")

    # Defina as credenciais e URLs do SharePoint
    sharepoint_site_url = "https://vallourec.sharepoint.com/teams/HOSHINKANRIPL/"
    username = tela.values['usuario']
    password = tela.values['senha']
    file_path = r'\\Srvoffice\pl\PLM\Coord_Movimentação_Equipamentos\10. Power BI\Logística Interna - Indicadores\Acompanhamento Diário\Base Carretas.xlsx'  # Caminho do arquivo que deseja fazer upload
    target_folder_url = r"/teams/HOSHINKANRIPL/Documentos%20Compartilhados/Gest%C3%A3o%20de%20Contrato/Base%20Carretas%20na%20Vaga"  # Caminho da pasta de destino no SharePoint

    # Autenticação
    ctx = ClientContext(sharepoint_site_url).with_credentials(UserCredential(username, password))

    # Lê o conteúdo do arquivo
    with open(file_path, 'rb') as content_file:
        file_content = content_file.read()

    # Faz o upload do arquivo
    target_folder = ctx.web.get_folder_by_server_relative_url(target_folder_url)
    name = os.path.basename(file_path)  # Nome do arquivo a ser salvo no SharePoint

    # Executa o upload
    target_file = target_folder.upload_file(name, file_content).execute_query()

    print(f"Arquivo '{name}' foi enviado com sucesso para {target_file.serverRelativeUrl}")

else:
    pass