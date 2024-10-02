from openpyxl import load_workbook
from relturno import TelaRelTurno
from PIL import Image
import openpyxl, os, glob, pyautogui, pyperclip
import PySimpleGUI as sg
from time import sleep
import pandas as pd
from datetime import date
from openpyxl.styles import Alignment, Side, Border
from matplotlib import pyplot as plt

user = os.path.expanduser('~')

arquivo = glob.glob(user + r'\Downloads\*VALLOURECADM*.xls')

center_align = Alignment(horizontal='center', vertical='center')
box = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

data_hoje = date.today().strftime("%d/%m/%Y")

barreiro_df = pd.read_excel(arquivo[0])

linha = barreiro_df[barreiro_df['Unnamed: 0'] == 'Veículos dentro das Áreas'].index[0]

barreiro_df.drop(barreiro_df.columns[10:17], axis=1, inplace=True)

barreiro_df.drop(barreiro_df.index[:linha+1], axis=0, inplace=True)

barreiro_df.columns = barreiro_df.iloc[0]

barreiro_df = barreiro_df[1:]

barreiro_df = barreiro_df.rename(columns={'Sub-grupo': 'Data', 'Grupo da Área': 'Identificador'})

barreiro_df['Data'] = data_hoje

barreiro_df['Identificador'] = 'Barreiro-Tradimaq'

barreiro_df = barreiro_df[barreiro_df['Categoria'].str.contains('OFICINA', na=False)]

x = barreiro_df['Tempo dentro'].str.split(':').str[0].astype(int)

x = x.sort_values(ascending=False)

y = barreiro_df['Nome do veículo']

fig, ax = plt.subplots(figsize =(16, 9))

ax.barh(y, x)

for s in ['top', 'bottom', 'left', 'right']:
    ax.spines[s].set_visible(False)

ax.grid(b = True, color ='grey',
        linestyle ='-.', linewidth = 0.5,
        alpha = 0.2)

ax.xaxis.set_ticks_position('none')
ax.yaxis.set_ticks_position('none')

ax.xaxis.set_tick_params(pad = 5)
ax.yaxis.set_tick_params(pad = 10)

for i in ax.patches:
    plt.text(i.get_width()+0.5, i.get_y()+0,
             str(round((i.get_width()), 2)),
             fontsize = 8,
             color ='black')

ax.set_title('Equipamentos no Canteiro',
             loc ='left', )

fig.text(0.9, 0.15, 'Shalkz', fontsize = 12,
         color ='grey', ha ='right', va ='bottom',
         alpha = 0.7)

plt.xlabel("Horas")

plt.show()

wb = load_workbook(r'\\srvoffice\PL\PLM\Coord_Movimentação_Equipamentos\4. Projetos\2023\Relatório de Turno\cobaia.xlsx')

ws = wb['RelatorioTurno']

tela = TelaRelTurno()

tela.Iniciar()

if tela.button in (sg.WINDOW_CLOSED, "Quit"):
    pass

else:

    ws['G2'].value = tela.values['responsavel']

    ws['I2'].value = tela.values['-CAL-']

    ws['K2'].value = tela.values['turno']

    ws['A7'].value = tela.values['infogeral']

    ws['A28'].value = tela.values['desvio_seg']

    ws['J28'].value = tela.values['desvio_meio']

    ws['A56'].value = tela.values['indisp']

    im = Image.open(tela.values['equip_canteiro'])

    newsize = (1320, 640)

    im1 = im.resize(newsize)

    im1.save(user + r'\Downloads\1.png', "PNG", optimize=True)    

    img = openpyxl.drawing.image.Image(user + r'\Downloads\1.png')

    img.anchor = 'A74'

    ws.add_image(img)

    im = Image.open(tela.values['vereagir'])

    newsize = (1320, 400)

    im1 = im.resize(newsize)

    im1.save(user + r'\Downloads\2.png', "PNG", optimize=True)

    img = openpyxl.drawing.image.Image(user + r'\Downloads\2.png')

    img.anchor = 'A109'    

    ws.add_image(img)

    aux = 2

    if tela.values['anexo_seg_1'] is not '':

        im = Image.open(tela.values['anexo_seg_1'])

        newsize = (198, 80)

        im1 = im.resize(newsize)

        aux = aux + 1

        im1.save(user + r'\Downloads\%s.png' % (aux), "PNG", optimize=True)

        img = openpyxl.drawing.image.Image(user + r'\Downloads\%s.png' % (aux))

        img.anchor = 'A47'

        ws.add_image(img)

    if tela.values['anexo_seg_2'] is not '':

        im = Image.open(tela.values['anexo_seg_2'])

        newsize = (215, 80)

        im1 = im.resize(newsize)

        aux = aux + 1

        im1.save(user + r'\Downloads\%s.png' % (aux), "PNG", optimize=True)

        img = openpyxl.drawing.image.Image(user + r'\Downloads\%s.png' % (aux))

        img.anchor = 'D47'

        ws.add_image(img)

    if tela.values['anexo_seg_3'] is not '':

        im = Image.open(tela.values['anexo_seg_3'])

        newsize = (295, 80)

        im1 = im.resize(newsize)

        aux = aux + 1

        im1.save(user + r'\Downloads\%s.png' % (aux), "PNG", optimize=True)

        img = openpyxl.drawing.image.Image(user + r'\Downloads\%s.png' % (aux))

        img.anchor = 'G47'

        ws.add_image(img)

    if tela.values['anexo_meio_1'] is not '':

        im = Image.open(tela.values['anexo_meio_1'])

        newsize = (168, 80)

        im1 = im.resize(newsize)

        aux = aux + 1

        im1.save(user + r'\Downloads\%s.png' % (aux), "PNG", optimize=True)

        img = openpyxl.drawing.image.Image(user + r'\Downloads\%s.png' % (aux))

        img.anchor = 'J47'

        ws.add_image(img)

    if tela.values['anexo_meio_2'] is not '':

        im = Image.open(tela.values['anexo_meio_2'])

        newsize = (140, 80)

        im1 = im.resize(newsize)

        aux = aux + 1

        im1.save(user + r'\Downloads\%s.png' % (aux), "PNG", optimize=True)

        img = openpyxl.drawing.image.Image(user + r'\Downloads\%s.png' % (aux))

        img.anchor = 'M47'

        ws.add_image(img)

    if tela.values['anexo_meio_3'] is not '':

        im = Image.open(tela.values['anexo_meio_3'])

        newsize = (309, 80)

        im1 = im.resize(newsize)

        aux = aux + 1

        im1.save(user + r'\Downloads\%s.png' % (aux), "PNG", optimize=True)

        img = openpyxl.drawing.image.Image(user + r'\Downloads\%s.png' % (aux))

        img.anchor = 'O47'

        ws.add_image(img)

    wb.save(r'\\srvoffice\PL\PLM\Coord_Movimentação_Equipamentos\4. Projetos\2023\Relatório de Turno\Relatório_%s_Turno %s.xlsx' % (tela.values['-CAL-'].replace("/","-"), tela.values['turno']))

    for i in range (1, aux+1):
        os.remove(user + r'\Downloads\%s.png' % (i))

    os.remove(arquivo[0])

    pyperclip.copy(r'\\srvoffice\PL\PLM\Coord_Movimentação_Equipamentos\4. Projetos\2023\Relatório de Turno\Relatório_%s_Turno %s.xlsx' % (tela.values['-CAL-'].replace("/","-"), tela.values['turno']))
    pyautogui.hotkey("win","r")
    sleep(1)
    pyautogui.hotkey("ctrl","v")
    sleep(1)
    pyautogui.press("enter")