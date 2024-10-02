from openpyxl import load_workbook
from datetime import datetime, timedelta
from suporte.Recursos import TelaCarretasso
import os, calendar

tela = TelaCarretasso()

tela.Iniciar()

if tela.button == "Gerar Dados":

    user = os.path.expanduser("~")

    def text_to_datetime(text):
        # Define o formato
        formato = "%d/%m/%Y %H:%M:%S"
        
        # Converte o texto para datetime
        dt = datetime.strptime(text, formato)
        
        return dt

    def texto_antes_primeiro_espaco(texto):
        # Encontra a posição do primeiro espaço
        posicao_espaco = texto.find(' ')
        
        # Retorna o texto antes do primeiro espaço, incluindo o espaço
        if posicao_espaco != -1:
            return texto[:posicao_espaco + 1]
        else:
            return texto  # Se não houver espaço, retorna o texto inteiro

    #Criar navegação nas worksheets
    wb = load_workbook(tela.values['doc'])

    ws = wb["Sheet1"]
    ws_dados = wb.create_sheet("Consolidado")
    ws_dados['A1'] = "Carreta"
    ws_dados['B1'] = "Tempo Parado"
    wsdados = wb["Consolidado"]

    #Deletar colunas desnecessárias
    ws.delete_cols(1,10)
    ws.delete_cols(2,5)
    ws.delete_cols(3,2)
    ws.delete_cols(5,2)
    ws.delete_cols(6,4)
    ws.delete_cols(7,8)
    ws.delete_cols(8,6)
    ws.delete_rows(1,3)

    #Pegar data final do mês vigente
    data = datetime.strptime(ws.cell(2,1).value[:10], "%d/%m/%Y")
    ultimo_dia = calendar.monthrange(data.year, data.month)[1]
    data_ultimo_dia = data.replace(day=ultimo_dia, hour=23, minute=59, second=59)

    for i in range(2, ws.max_row+1):
        ws.cell(i,1).value = text_to_datetime(ws.cell(i,1).value)
        ws.cell(i,2).value = text_to_datetime(ws.cell(i,2).value)

    for i in range(2, ws.max_row+1):
        if ws.cell(i,5).value == "CANTEIRO TRADIMAQ 1" and ws.cell(i,6).value == "---" or ws.cell(i,5).value == "CANTEIRO TRADIMAQ 2" and ws.cell(i,6).value == "---":
            tempo_inicio = data_ultimo_dia
            # print(ws.cell(i,3).value)
            estado = False
            for j in range(2, ws.max_row+1):
                if texto_antes_primeiro_espaco(ws.cell(i, 3).value) in ws.cell(j, 3).value and ws.cell(i, 2).value < ws.cell(j, 1).value:
                    if tempo_inicio > ws.cell(j,1).value and tempo_inicio > ws.cell(i,2).value:
                        tempo_inicio = ws.cell(j,1).value
                        estado = True
            if estado:
                tempo_parado = tempo_inicio - ws.cell(i,2).value
                wsdados.cell(wsdados.max_row+1,1).value = ws.cell(i,3).value 
                wsdados.cell(wsdados.max_row,2).value = tempo_parado
            if not estado:
                tempo_parado = data_ultimo_dia - ws.cell(i,2).value
                wsdados.cell(wsdados.max_row+1,1).value = ws.cell(i,3).value 
                wsdados.cell(wsdados.max_row,2).value = tempo_parado

    wb.save(user + r"\Downloads\Resultado (Tempos).xlsx")

else:
    pass