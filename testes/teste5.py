from openpyxl import load_workbook
import os

user = os.path.expanduser('~')

wb = load_workbook(filename=(user + '\\Downloads\\valor.xlsx'))

ws = wb['Plan1']

ws.cell(1, 0).value = 5

wb.save