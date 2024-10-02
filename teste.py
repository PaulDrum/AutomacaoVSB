from openpyxl import load_workbook
import os

user = os.path.expanduser('~')

wb = load_workbook(filename=(user + '\\Downloads\\Relatorio_viagens_carretinhas_ruckers.xlsx'))

origemws = wb['Sheet1']

def remove_unnecessary_columns(ws):
        # Deletar colunas de A a N (1 a 14)
        ws.delete_cols(1, 14)
        
        # Deletar coluna B (agora é a coluna A após a remoção anterior)
        ws.delete_cols(2, 1)  # Deleta a nova coluna B

        # Deletar colunas C e D (agora são as colunas B e C após as remoções anteriores)
        ws.delete_cols(3, 2)  # Deleta as novas colunas B e C

        # Deletar colunas D ao K (agora são as colunas B ao J após as remoções anteriores)
        ws.delete_cols(4, 8)  # Deleta as novas colunas B até J

        # Deletar colunas E ao L (agora são as colunas B ao K após as remoções anteriores)
        ws.delete_cols(5, 8)  # Deleta as novas colunas B até H

        # Deletar tudo após a coluna E (agora é a coluna A após todas as remoções anteriores)
        max_col = ws.max_column
        if max_col > 5:  # Se houver mais de 5 colunas
            ws.delete_cols(6, max_col - 5)  # Deleta da coluna F até a última

# Remover colunas desnecessárias na planilha de origem
remove_unnecessary_columns(origemws)

wb.save(filename=(user + '\\Downloads\\Teste.xlsx'))
