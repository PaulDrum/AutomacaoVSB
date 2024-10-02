from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from suporte.Recursos import AcesseMeuSite, TelaJornadas
import pyautogui
from openpyxl import load_workbook

tela = TelaJornadas()

tela

if tela.values['6x1'] or tela.values['12x36'] == True:

    conta = "VALLOUREC-GESTAO"

    Turnos = ['Turno 1', 'Turno 2', 'Turno 3', '12 x 36 Dia', '12 x 36 Noite']

    intervalo = 60

    # chrome_options = webdriver.EdgeOptions()
    # chrome_options.add_argument('--start-maximized')
    # chrome_options.add_argument("--disable-infobars")
    # chrome_options.add_experimental_option("useAutomationExtension", False)
    # chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    # chrome_options.add_argument("--window-position=0,0")
    # chrome_options.add_experimental_option("detach", True)

    navegador = webdriver.Edge(executable_path=r'C:\Users\s-Lucas.SAraujo\AppData\Local\Programs\Python\Python36-32\msedgedriver.exe')

    sleep(5)
    
    navegador.switch_to.window(navegador.window_handles[0])

    navegador.get("https://web.pdo-vallourec.com.br/#/")

    navegador.maximize_window

    sleep(15)

    AcesseMeuSite.Logar(navegador, conta, tela.values['usuario'], tela.values['senha'])

    sleep(15)

    navegador.find_element(By.XPATH, "//*[contains(text(),'Atendimento')]").click()

    sleep(1)

    navegador.find_element(By.XPATH, "//*[contains(text(),'Programação de Jornada')]").click()

    sleep(5)

    wb = load_workbook(tela.values['doc'])

if tela.values['6x1'] == True:

    ws = wb['Barreiro']

    max_row_ws = ws.max_row

    MatrizTurnos = ["T1", "T2", "T3", "T1", "T2", "T3"]

    for i in range(2, max_row_ws+1):
        for j in range(5, 11):
            if ws.cell(i, j).value is not None:
                MatrizTurnos[j - 5] = ws.cell(i, j).value
            else:
                MatrizTurnos[j - 5] = 0

        for n in range(0, 6):

            if n == 3 and MatrizTurnos[0] != 0:
                continue

            if n == 4 and MatrizTurnos[1] != 0:
                continue

            if n == 5 and MatrizTurnos[2] != 0:
                continue

            if MatrizTurnos[n] != 0:
                    
                qtdJornadas = MatrizTurnos[n]

                navegador.find_element(By.XPATH, "//*[contains(text(),'Nova programação de jornada')]").click()
                sleep(1)   

                # testando datas
                navegador.find_element(By.XPATH, '//*[@placeholder="Data inicial"]').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                sleep(0.2)
                navegador.find_element(By.XPATH, '//*[@placeholder="Data inicial"]').send_keys(tela.values["-CAL-"])
                sleep(0.2)
                navegador.find_element(By.XPATH, '//*[@placeholder="Data final"]').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                sleep(0.2)
                navegador.find_element(By.XPATH, '//*[@placeholder="Data final"]').send_keys(tela.values["-CAL2-"])
                sleep(0.2)

                if ws.cell(i, 3).value == "SEG/SAB":
                    navegador.find_element(By.XPATH, "//*[contains(text(),'Dias da semana')]").click()
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[4]/div/div/mat-checkbox/label/span').click()
                    sleep(0.2)
                    navegador.find_element(By.XPATH, "/html/body/div[2]/div[4]/div/div/mat-option[1]/span").click()
                    sleep(0.2)
                    pyautogui.doubleClick(x=658, y=325)
                    sleep(0.2)
                
                elif ws.cell(i, 3).value == "SEG/SEX":
                    navegador.find_element(By.XPATH, "//*[contains(text(),'Dias da semana')]").click()
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[4]/div/div/mat-checkbox/label/span').click()
                    sleep(0.2)
                    navegador.find_element(By.XPATH, "/html/body/div[2]/div[4]/div/div/mat-option[1]/span").click()
                    sleep(0.2)
                    navegador.find_element(By.XPATH, "/html/body/div[2]/div[4]/div/div/mat-option[7]/span").click()
                    sleep(0.2)
                    pyautogui.doubleClick(x=658, y=325)
                    sleep(0.2)

                elif ws.cell(i, 3).value == "DOM/SEXTA":
                    navegador.find_element(By.XPATH, "//*[contains(text(),'Dias da semana')]").click()
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[4]/div/div/mat-checkbox/label/span').click()
                    sleep(0.2)
                    navegador.find_element(By.XPATH, "/html/body/div[2]/div[4]/div/div/mat-option[7]/span").click()
                    sleep(0.2)
                    pyautogui.doubleClick(x=658, y=325)
                    sleep(0.2)

                else:
                    navegador.find_element(By.XPATH, "//*[contains(text(),'Dias da semana')]").click()
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[4]/div/div/mat-checkbox/label/span').click()
                    sleep(0.2)
                    pyautogui.doubleClick(x=658, y=325)
                    sleep(0.2)

                classe = ws.cell(i, 1).value
                
                navegador.find_element(By.XPATH, "//*[contains(text(), '--- Selecione a classe ---')]").click()
                sleep(0.2)
                navegador.find_element(By.XPATH, f"//*[contains(text(),'{classe}')]").click()
                sleep(0.2)
                
                area = ws.cell(i, 2).value

                navegador.find_element(By.XPATH, "//*[contains(text(), 'AC04')]").click()
                sleep(0.2)
                navegador.find_element(By.XPATH, f"//*[text()='{area}']").click()
                sleep(0.2)

                if ws.cell(i, 4).value == "SIM":

                    navegador.find_element(By.XPATH, '//*[@placeholder="Tempo de intervalo (em minutos)"]').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '//*[@placeholder="Tempo de intervalo (em minutos)"]').send_keys(intervalo)
                    sleep(0.2)

                if qtdJornadas != 1:
                    #testando qtd_jornadas
                    navegador.find_element(By.XPATH, '//*[@placeholder="Quantidade de jornadas"]').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '//*[@placeholder="Quantidade de jornadas"]').send_keys(qtdJornadas)
                    sleep(0.2)

                if n == 0 or n == 3:
                        
                    #testando turno
                    navegador.find_element(By.XPATH, "//*[contains(text(), '--- Selecione o turno ---')]").click()
                    sleep(0.2)
                    navegador.find_element(By.XPATH, f"//*[contains(text(), '{Turnos[0]}')]").click()
                    sleep(0.2)

                    #testando horas
                    navegador.find_element(By.XPATH, '//*[@placeholder="Hora final"]').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '//*[@placeholder="Hora final"]').send_keys(tela.values["t1fim"])
                    sleep(0.2)

                    navegador.find_element(By.XPATH, '//*[contains(text(), "Salvar")]').click()

                elif n == 1 or n == 4:

                    #testando turno
                    navegador.find_element(By.XPATH, "//*[contains(text(), '--- Selecione o turno ---')]").click()
                    sleep(0.2)
                    navegador.find_element(By.XPATH, f"//*[contains(text(), '{Turnos[1]}')]").click()
                    sleep(0.2)

                    #testando horas
                    navegador.find_element(By.XPATH, '//*[@placeholder="Hora inicial"]').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '//*[@placeholder="Hora inicial"]').send_keys(tela.values["t2inicio"])
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '//*[@placeholder="Hora final"]').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '//*[@placeholder="Hora final"]').send_keys(tela.values["t2fim"])
                    sleep(0.2)
                    
                    navegador.find_element(By.XPATH, '//*[contains(text(), "Salvar")]').click()

                else:

                    #testando turno
                    navegador.find_element(By.XPATH, "//*[contains(text(), '--- Selecione o turno ---')]").click()
                    sleep(0.2)
                    navegador.find_element(By.XPATH, f"//*[contains(text(), '{Turnos[2]}')]").click()
                    sleep(0.2)

                    #testando horas
                    navegador.find_element(By.XPATH, '//*[@placeholder="Hora inicial"]').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '//*[@placeholder="Hora inicial"]').send_keys(tela.values["t3inicio"])
                    sleep(0.2)

                    navegador.find_element(By.XPATH, '//*[contains(text(), "Salvar")]').click()

                sleep(2)         

if tela.values['12x36'] == True:   

    ws = wb['Barreiro 12x36']

    max_row_ws = ws.max_row

    MatrizTurnos = ['S1', 'S2', 'D1', 'D2']

    for i in range(2, max_row_ws+1):

        for j in range(5, 9):
            if ws.cell(i, j).value is not None:
                MatrizTurnos[j - 5] = ws.cell(i, j).value
            else:
                MatrizTurnos[j - 5] = 0

        for n in range(0, 4):

            if n == 2 and  MatrizTurnos[0] != 0:
                continue

            if n == 3 and  MatrizTurnos[1] != 0:
                continue

            if MatrizTurnos[n] != 0:

                if n == 1 or n == 3:

                    qtdJornadas = MatrizTurnos[n]

                    navegador.find_element(By.XPATH, "//*[contains(text(),'Nova programação de jornada')]").click()
                    sleep(1)   

                    # testando datas
                    navegador.find_element(By.XPATH, '//*[@placeholder="Data inicial"]').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '//*[@placeholder="Data inicial"]').send_keys(tela.values["-CAL-"])
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '//*[@placeholder="Data final"]').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '//*[@placeholder="Data final"]').send_keys(tela.values["-CAL2-"])
                    sleep(0.2)

                    navegador.find_element(By.XPATH, "//*[contains(text(),'Dias da semana')]").click()
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[4]/div/div/mat-checkbox/label/span').click()
                    sleep(0.2)
                    pyautogui.doubleClick(x=658, y=325)
                    sleep(0.2)

                    classe = ws.cell(i, 1).value
                    
                    navegador.find_element(By.XPATH, "//*[contains(text(), '--- Selecione a classe ---')]").click()
                    sleep(0.2)
                    navegador.find_element(By.XPATH, f"//*[contains(text(),'{classe}')]").click()
                    sleep(0.2)
                    
                    area = ws.cell(i, 2).value

                    navegador.find_element(By.XPATH, "//*[contains(text(), 'AC04')]").click()
                    sleep(0.2)
                    navegador.find_element(By.XPATH, f"//*[text()='{area}']").click()
                    sleep(0.2)

                    if ws.cell(i, 4).value == "SIM":
                        navegador.find_element(By.XPATH, '//*[@placeholder="Tempo de intervalo (em minutos)"]').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                        sleep(0.2)
                        navegador.find_element(By.XPATH, '//*[@placeholder="Tempo de intervalo (em minutos)"]').send_keys(intervalo)
                        sleep(0.2)

                    if qtdJornadas != 1:
                        #testando qtd_jornadas
                        navegador.find_element(By.XPATH, '//*[@placeholder="Quantidade de jornadas"]').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                        sleep(0.2)
                        navegador.find_element(By.XPATH, '//*[@placeholder="Quantidade de jornadas"]').send_keys(qtdJornadas)
                        sleep(0.2)
    
                    #testando turno
                    navegador.find_element(By.XPATH, "//*[contains(text(), '--- Selecione o turno ---')]").click()
                    sleep(0.2)
                    navegador.find_element(By.XPATH, f"//*[contains(text(), '{Turnos[3]}')]").click()
                    sleep(0.2)

                    #testando horas
                    navegador.find_element(By.XPATH, '//*[@placeholder="Hora inicial"]').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '//*[@placeholder="Hora inicial"]').send_keys(tela.values["12x36dinicio"])
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '//*[@placeholder="Hora final"]').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '//*[@placeholder="Hora final"]').send_keys(tela.values["12x36dfim"])
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '//*[contains(text(), "Salvar")]').click()

                    sleep(2) 

                else:    
                                
                    qtdJornadas = MatrizTurnos[n]

                    navegador.find_element(By.XPATH, "//*[contains(text(),'Nova programação de jornada')]").click()
                    sleep(1)   

                    # testando datas
                    navegador.find_element(By.XPATH, '//*[@placeholder="Data inicial"]').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '//*[@placeholder="Data inicial"]').send_keys(tela.values["-CAL-"])
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '//*[@placeholder="Data final"]').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '//*[@placeholder="Data final"]').send_keys(tela.values["-CAL2-"])
                    sleep(0.2)

                    navegador.find_element(By.XPATH, "//*[contains(text(),'Dias da semana')]").click()
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[4]/div/div/mat-checkbox/label/span').click()
                    sleep(0.2)
                    pyautogui.doubleClick(x=658, y=325)
                    sleep(0.2)

                    classe = ws.cell(i, 1).value
                    
                    navegador.find_element(By.XPATH, "//*[contains(text(), '--- Selecione a classe ---')]").click()
                    sleep(0.2)
                    navegador.find_element(By.XPATH, f"//*[contains(text(),'{classe}')]").click()
                    sleep(0.2)
                    
                    area = ws.cell(i, 2).value

                    navegador.find_element(By.XPATH, "//*[contains(text(), 'AC04')]").click()
                    sleep(0.2)
                    navegador.find_element(By.XPATH, f"//*[text()='{area}']").click()
                    sleep(0.2)

                    if qtdJornadas != 1:
                        #testando qtd_jornadas
                        navegador.find_element(By.XPATH, '//*[@placeholder="Quantidade de jornadas"]').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                        sleep(0.2)
                        navegador.find_element(By.XPATH, '//*[@placeholder="Quantidade de jornadas"]').send_keys(qtdJornadas)
                        sleep(0.2)
    
                    #testando turno
                    navegador.find_element(By.XPATH, "//*[contains(text(), '--- Selecione o turno ---')]").click()
                    sleep(0.2)
                    navegador.find_element(By.XPATH, f"//*[contains(text(), '{Turnos[4]}')]").click()
                    sleep(0.2)

                    #testando horas
                    navegador.find_element(By.XPATH, '//*[@placeholder="Hora inicial"]').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '//*[@placeholder="Hora inicial"]').send_keys(tela.values["12x36ninicio"])
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '//*[contains(text(), "Salvar")]').click()

                    sleep(2)   

                    navegador.find_element(By.XPATH, "//*[contains(text(),'Nova programação de jornada')]").click()
                    sleep(1)   

                    # testando datas
                    navegador.find_element(By.XPATH, '//*[@placeholder="Data inicial"]').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '//*[@placeholder="Data inicial"]').send_keys(tela.values["-CAL-"])
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '//*[@placeholder="Data final"]').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '//*[@placeholder="Data final"]').send_keys(tela.values["-CAL2-"])
                    sleep(0.2)

                    navegador.find_element(By.XPATH, "//*[contains(text(),'Dias da semana')]").click()
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '/html/body/div[2]/div[4]/div/div/mat-checkbox/label/span').click()
                    sleep(0.2)
                    pyautogui.doubleClick(x=658, y=325)
                    sleep(0.2)

                    classe = ws.cell(i, 1).value
                    
                    navegador.find_element(By.XPATH, "//*[contains(text(), '--- Selecione a classe ---')]").click()
                    sleep(0.2)
                    navegador.find_element(By.XPATH, f"//*[contains(text(),'{classe}')]").click()
                    sleep(0.2)
                    
                    area = ws.cell(i, 2).value

                    navegador.find_element(By.XPATH, "//*[contains(text(), 'AC04')]").click()
                    sleep(0.2)
                    navegador.find_element(By.XPATH, f"//*[text()='{area}']").click()
                    sleep(0.2)

                    if ws.cell(i, 4).value == "SIM":
                        navegador.find_element(By.XPATH, '//*[@placeholder="Tempo de intervalo (em minutos)"]').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                        sleep(0.2)
                        navegador.find_element(By.XPATH, '//*[@placeholder="Tempo de intervalo (em minutos)"]').send_keys(intervalo)
                        sleep(0.2)

                    if qtdJornadas != 1:
                        #testando qtd_jornadas
                        navegador.find_element(By.XPATH, '//*[@placeholder="Quantidade de jornadas"]').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                        sleep(0.2)
                        navegador.find_element(By.XPATH, '//*[@placeholder="Quantidade de jornadas"]').send_keys(qtdJornadas)
                        sleep(0.2)
    
                    #testando turno
                    navegador.find_element(By.XPATH, "//*[contains(text(), '--- Selecione o turno ---')]").click()
                    sleep(0.2)
                    navegador.find_element(By.XPATH, f"//*[contains(text(), '{Turnos[4]}')]").click()
                    sleep(0.2)

                    #testando horas
                    navegador.find_element(By.XPATH, '//*[@placeholder="Hora final"]').send_keys(Keys.LEFT_CONTROL + 'a', Keys.DELETE)
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '//*[@placeholder="Hora final"]').send_keys(tela.values["12x36nfim"])
                    sleep(0.2)
                    navegador.find_element(By.XPATH, '//*[contains(text(), "Salvar")]').click()

                    sleep(2)           