import openpyxl, csv

wb = openpyxl.Workbook()
ws = wb.active

with open(r"C:\Users\s-Lucas.SAraujo\Downloads\Alerta de Qualidade.csv",encoding='utf-8') as f:
    reader = csv.reader(f, delimiter=',')
    for row in reader:
        ws.append(row)

wb.save(r"C:\Users\s-Lucas.SAraujo\Downloads\Alerta de Qualidade.xlsx")

wb = openpyxl.load_workbook(r"C:\Users\s-Lucas.SAraujo\Downloads\Alerta de Qualidade.xlsx")
wb2 = openpyxl.load_workbook(r"C:\Users\s-Lucas.SAraujo\Downloads\Teste - Copia.xlsx")

ws = wb['Sheet']
ws2 = wb2['Sheet']

max_row = ws.max_row
max_row2 = ws2.max_row

max_column = ws.max_column

logistica = "UC - Central Supply Chain"

i = 2

while i <= max_row:
    if ws.cell(i, 5).value is None:
        break
    if ws.cell(i, 5).value != logistica:
        ws.delete_rows(i, 1)
    else:
        i+=1

max_row = ws.max_row

# Parte de remover duplicatas

# i = 2

# while i <= max_row:
#     for j in range(2, max_row2+1):
#         if ws.cell(i, 2).value == ws2.cell(j, 2).value:
#             ws.delete_rows(i, 1)
#         else:
#             i+=1

dados_pessoais = [
    ["RESENDE Leone", "DPA", "T1"],
    ["OLIVEIRA Marcio Cristiano", "DPA", "T3"],
    ["MIRANDA Rodrigo", "Plat Sul", "T3"],
    ["GUEDES Reginaldo", "Plat Sul", "T2"],
    ["SANTOS Mauro", "Plat Sul", "T3"],
    ["SERRA Pedro Henrique", "DPA", "T2"],
    ["ANTUNES Cristiano", "Plat Sul", "T1"],
    ["RIBEIRO Juarez", "Plat Sul", "T1"],
    ["COUTINHO Tiago", "DPA", "T1"],
    ["PAULA Marcio", "Plat Sul", "T2"],
    ["SILVEIRA Jucimar", "DPA", "T3"],
    ["DANTAS Pedro", "Plat Sul", "T2"],
    ["BARBOSA Adriano", "Plat Sul", "T2"],
    ["VENANCIO Julio Cesar", "Plat Sul", "T2"],
    ["DIAS SILVA Lucas", "DPA", "T1"],
    ["CASTRO William", "Plat Sul", "T3"],
    ["BENJAMIM Leandro", "DPA", "T2"],
    ["DALFIOR João", "DPA", "T3"],
    ["FRAGA Marcelo", "Plat Sul", "T2"],
    ["COSTA Elias", "Plat Sul", "T2"],
    ["LIMA Carla Cristina", "DPA", "T2"],
    ["LOPES Rodrigo", "Plat Sul", "T2"],
    ["PIMENTEL Tiago", "DPA", "T3"],
    ["MACEDO Wendel", "Plat Sul", "T3"],
    ["XAVIER Jose", "DPA", "T2"],
    ["DIAS Andre", "DPA", "T2"],
    ["CRUZ Ricardo", "DPA", "T2"],
    ["SOUZA Warlei", "Plat Sul", "T2"],
    ["SANTOS Heledy", "DPA", "T2"],
    ["MONTEIRO Robert", "Plat Sul", "T1"],
    ["ROCHA Devanir", "Plat Sul", "T2"]
]

ws.cell(1, 49).value == 'Área'
ws.cell(1, 50).value == 'Turno'

for i in range(2, max_row+1):
    for name in dados_pessoais:
        if ws.cell(i, 1).value == name[0]:
            ws.cell(i, 49).value == name[1]
            ws.cell(i, 50).value == name[2]

wb.save(r"C:\Users\s-Lucas.SAraujo\Downloads\Teste.xlsx")