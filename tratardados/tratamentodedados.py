import pandas as pd
import os
from datetime import date
from openpyxl.styles import Alignment, Side, Border
import xlsxwriter
from openpyxl import load_workbook
from time import sleep
import datetime

center_align = Alignment(horizontal='center', vertical='center')
box = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

data_hoje = date.today().strftime("%d/%m/%Y")

user = os.path.expanduser('~')

barreiro_df = pd.read_excel(user + r'\Downloads\barreiro.xls')
jeceaba_df = pd.read_excel(user + r'\Downloads\jeceaba.xls')

linha = barreiro_df[barreiro_df['Unnamed: 0'] == 'Veículos dentro das Áreas'].index[0]

barreiro_df.drop(barreiro_df.columns[10:17], axis=1, inplace=True)

barreiro_df.drop(barreiro_df.index[:linha+1], axis=0, inplace=True)

barreiro_df.columns = barreiro_df.iloc[0]

barreiro_df = barreiro_df[1:]

barreiro_df = barreiro_df.rename(columns={'Sub-grupo': 'Data', 'Grupo da Área': 'Identificador'})

barreiro_df['Data'] = data_hoje

barreiro_df['Identificador'] = 'Barreiro-Tradimaq'

barreiro_df = barreiro_df[barreiro_df['Categoria'].str.contains('OFICINA', na=False)]

linha = jeceaba_df[jeceaba_df['Unnamed: 0'] == 'Veículos dentro das Áreas'].index[0]

jeceaba_df.drop(jeceaba_df.columns[10:17], axis=1, inplace=True)

jeceaba_df.drop(jeceaba_df.index[:linha+1], axis=0, inplace=True)

jeceaba_df.columns = jeceaba_df.iloc[0]

jeceaba_df = jeceaba_df[1:]

jeceaba_df = jeceaba_df.rename(columns={'Sub-grupo': 'Data', 'Grupo da Área': 'Identificador'})

jeceaba_df['Data'] = data_hoje

jeceaba_df['Identificador'] = 'Jeceaba-Tradimaq'

frames = [barreiro_df, jeceaba_df]

result = pd.concat(frames)

result = result.reset_index(drop=True)

writer = pd.ExcelWriter(user + r'\Downloads\Equipamentos em Manutenção.xlsx', engine='xlsxwriter')
result.to_excel(writer, sheet_name='Sheet1')
workbook = writer.book
worksheet = writer.sheets['Sheet1']

border_fmt = workbook.add_format({'bottom':1, 'top':1, 'left':1, 'right':1})
worksheet.conditional_format(xlsxwriter.utility.xl_range(0, 0, len(result), len(result.columns)), {'type': 'no_errors', 'format': border_fmt})
writer.save()

wb = load_workbook(filename=(user + '\\Downloads\\Equipamentos em Manutenção.xlsx'))

ws = wb['Sheet1']

ws.delete_cols(1, 1)

for i in range(2, ws.max_row + 1):
    ws.cell(i, 2).value = datetime.datetime.strptime(ws.cell(i, 2).value, '%d/%m/%Y').date()

os.remove(user + '\\Downloads\\barreiro.xls')
os.remove(user + '\\Downloads\\jeceaba.xls')

wb.save(filename=(user + '\\Downloads\\Equipamentos em Manutenção.xlsx'))