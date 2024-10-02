from time import sleep
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from Recursos import AcesseMeuSiteAnalitico, Loop
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Side, Border
import os
import pyperclip
import pyautogui

user = os.path.expanduser('~')

center_align = Alignment(horizontal='center', vertical='center')
box = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

usuario = "VITOR.VALADARES"
senha = "Vitor@2022"
conta_barreiro = "VALLOUREC-GESTAO"
conta_jeceaba = "VALLOUREC-JECEABA"

navegador = webdriver.Chrome()

navegador.get('https://analitico-vallourec.com.br/login/')

#yesterday = date.today() - timedelta(days=1)
#d1 = yesterday.strftime("%d/%m/%Y")

action = ActionChains(navegador)

AcesseMeuSiteAnalitico.Logar(navegador, usuario.lower(), senha)

sleep(20)

navegador.find_element_by_xpath('/html/body/nav/div[2]/div/ul/li[1]').click()

sleep(5)

navegador.find_element_by_xpath('/html/body/nav/div[2]/div/ul/li[1]').click()

sleep(5)

pyperclip.copy('13/12/2022')

sleep(1)

navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/div/p[1]/input').send_keys(Keys.LEFT_CONTROL + "a")

sleep(1)

pyautogui.hotkey("ctrl","v")

sleep(5)

pyperclip.copy('13/12/2022')

sleep(1)

navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/div/p[2]/input').send_keys(Keys.LEFT_CONTROL + "a")

sleep(1)

pyautogui.hotkey("ctrl","v")

sleep(5)

navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[1]').click()

sleep(3)

navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[21]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[22]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[23]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[24]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[25]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[33]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[34]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[35]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[36]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[37]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[38]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[39]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[40]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[41]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[42]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[43]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[44]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[45]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[46]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[47]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[48]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[49]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[50]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[51]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[52]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[53]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[54]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[55]/input').click()
navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[56]/input').click()

sleep(5)

navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/button[3]').click()

sleep(120)

source = navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/button[2]')

sleep(1)

action.double_click(source).perform()

sleep(1)

for i in range(128, 153):
    navegador.find_element_by_xpath(f'/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/div/div[{i}]/input').click()

sleep(1)

navegador.find_element_by_xpath('/html/body/div[2]/div[2]/div/div/div/div/uib-accordion/div/div[27]/div[2]/div/button[3]').click()

sleep(120)

navegador.close()

wb = load_workbook(filename = user + r'\Downloads\Relatorio_permanencia_carreta_tempo.xlsx')
wb2 = load_workbook(filename = user + r'\Downloads\Relatorio_permanencia_carreta_tempo (1).xlsx')
wb3 = load_workbook(r'\\srvoffice\PL\PLM\Coord_Movimentação_Equipamentos\10. Power BI\Logística Interna - Indicadores\Acompanhamento Diário\Relatorio_permanencia_carreta_tempo.xlsx')

ws = wb['Sheet1']
ws2 = wb2['Sheet1']
ws3 = wb3['Sheet1']

max_row2 = ws2.max_row
max_row3 = ws3.max_row
max_row_ws = ws.max_row
max_col3 = ws3.max_column

Loop.Unico(4, 5, max_row2, ws2, center_align, box, 'PLATSUL')

Loop.Unico(4, 5, max_row_ws, ws, center_align, box, 'DPA')
    
max_col2 = ws2.max_column
max_col = ws.max_column
    
Loop.Duplo2(5, max_row2, 1, max_col2, max_row_ws, ws2, ws, center_align, box)
max_row2 = ws2.max_row

Loop.Duplo2(5, max_row3, 1, max_col3, max_row2, ws3, ws2, center_align, box)
max_row3 = ws3.max_row

wb3.save(filename = r'\\srvoffice\PL\PLM\Coord_Movimentação_Equipamentos\10. Power BI\Logística Interna - Indicadores\Acompanhamento Diário\Relatorio_permanencia_carreta_tempo.xlsx')

pyperclip.copy(r'\\srvoffice\PL\PLM\Coord_Movimentação_Equipamentos\10. Power BI\Logística Interna - Indicadores\Acompanhamento Diário\Carretas.pbix')
pyautogui.hotkey("win","r")
sleep(1)
pyautogui.hotkey("ctrl","v")
sleep(1)
pyautogui.press("enter")

os.remove(user + r'\Downloads\Relatorio_permanencia_carreta_tempo (1).xlsx')
os.remove(user + r'\Downloads\Relatorio_permanencia_carreta_tempo.xlsx')